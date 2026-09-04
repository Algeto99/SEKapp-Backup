import ast
import base64
import hashlib
import hmac
import json
import logging
import math
import os
import re
import smtplib
import ssl
from datetime import timedelta, datetime, date, time
from decimal import Decimal
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps
from io import BytesIO

import psycopg2
from psycopg2 import extras, errors as pg_errors
from flask import Blueprint, current_app, render_template, request, jsonify, redirect, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt, unset_jwt_cookies
import google.auth
from google.auth.transport import requests as google_auth_requests
from google.cloud import storage
from google.oauth2 import service_account

from markupsafe import escape
from db import get_db_connection

# PDF generation imports
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except OSError:
    print("WARNING: WeasyPrint could not be loaded. PDF generation will not work locally.")
    WEASYPRINT_AVAILABLE = False
    HTML = None

viewer_bp = Blueprint('viewer', __name__)
app_logger = logging.getLogger(__name__)

# Cached so a record with ten photos does not make ten metadata-server round trips.
_signing_credentials = None


def _signing_kwargs():
    """Extra arguments that let generate_signed_url actually sign on Cloud Run.

    The ambient credentials there come from the metadata server and carry only a
    bearer token, no private key, so the storage client cannot produce a v4
    signature locally — it raises "you need a private key to sign credentials" and
    the caller below falls back to an unsigned URL, which a private bucket answers
    with 403. Handing the client a service account email plus a live access token
    makes it sign through the IAM SignBlob API instead.

    Requires the runtime service account to hold roles/iam.serviceAccountTokenCreator
    on itself, and the IAM Service Account Credentials API to be enabled. Without
    that grant signing still fails, just with a 403 from iamcredentials instead.

    A JSON-key service account (local development) holds its own private key and
    signs locally, so it gets nothing added.
    """
    global _signing_credentials

    if _signing_credentials is None:
        _signing_credentials, _ = google.auth.default()

    if isinstance(_signing_credentials, service_account.Credentials):
        return {}

    # service_account_email is only filled in from the metadata server on the first
    # refresh, and SignBlob rejects a stale token, so both need a live credential.
    if not _signing_credentials.valid:
        _signing_credentials.refresh(google_auth_requests.Request())

    email = getattr(_signing_credentials, 'service_account_email', None)
    if not email or email == 'default' or not _signing_credentials.token:
        return {}

    return {
        'service_account_email': email,
        'access_token': _signing_credentials.token,
    }


def generate_signed_url(gcs_url):
    """Generates a v4 signed URL for a GCS blob."""
    try:
        if not gcs_url or 'storage.googleapis.com' not in gcs_url:
            return gcs_url
        
        # Extract bucket and blob name
        # Format: https://storage.googleapis.com/BUCKET_NAME/BLOB_NAME
        parts = gcs_url.replace("https://storage.googleapis.com/", "").split('/', 1)
        if len(parts) != 2:
            return gcs_url
            
        bucket_name, blob_name = parts
        
        # Explicitly use default credentials which should pick up the service account
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        
        url = blob.generate_signed_url(
            version="v4",
            expiration=timedelta(minutes=60), # 1 hour expiration
            method="GET",
            **_signing_kwargs()
        )
        app_logger.info(f"Generated signed URL for {blob_name}")
        return url
    except Exception as e:
        app_logger.error(f"Error generating signed URL for {gcs_url}: {e}", exc_info=True)
        return gcs_url


def _es_columna_de_imagen(col_name):
    """
    Columnas cuyo valor es una URL de GCS que hay que firmar para que se vea.

    Los objetos del bucket son privados (upload_file_to_gcs no los publica), así
    que una URL cruda no renderiza. Hasta ahora solo se firmaban las firmas, de
    modo que foto_evidencia_url y las cuatro fotos del pre-operacional llegaban
    sin firmar al detalle del registro.
    """
    lower = str(col_name or '').lower()
    return any(t in lower for t in ('firma', 'foto', 'evidencia', 'imagen', 'photo'))


def _make_media_token(gcs_base_url):
    """Return a URL-safe token encoding a GCS base URL with an HMAC signature."""
    secret = current_app.config.get('SECRET_KEY', '').encode()
    encoded = base64.urlsafe_b64encode(gcs_base_url.encode()).decode().rstrip('=')
    sig = hmac.new(secret, gcs_base_url.encode(), hashlib.sha256).hexdigest()[:24]
    return f"{encoded}.{sig}"


def _media_proxy_url(url):
    """Convert a GCS URL into a proxy URL that hides bucket and path details."""
    gcs_base = url.split('?')[0]
    token = _make_media_token(gcs_base)
    host = request.host_url.rstrip('/')
    return f"{host}/api/media?f={token}"


@viewer_bp.route('/api/media')
@jwt_required()
def serve_media():
    """Proxy GCS files through a signed-token redirect to hide bucket details."""
    token = request.args.get('f', '')
    try:
        parts = token.rsplit('.', 1)
        if len(parts) != 2:
            return 'Invalid request', 400
        encoded, sig = parts
        padded = encoded + '=' * (4 - len(encoded) % 4)
        gcs_url = base64.urlsafe_b64decode(padded).decode()
        if not gcs_url.startswith('https://storage.googleapis.com/'):
            return 'Invalid request', 400
        secret = current_app.config.get('SECRET_KEY', '').encode()
        expected = hmac.new(secret, gcs_url.encode(), hashlib.sha256).hexdigest()[:24]
        if not hmac.compare_digest(sig, expected):
            return 'Forbidden', 403
        signed_url = generate_signed_url(gcs_url)
        return redirect(signed_url, 302)
    except Exception as e:
        app_logger.warning(f"Media proxy error: {e}")
        return 'Invalid request', 400


_SCHEMA_CACHE = {}


def _get_table_columns(cur, table_name):
    if table_name not in _SCHEMA_CACHE:
        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = %s
        """, (table_name,))
        _SCHEMA_CACHE[table_name] = {row[0] for row in cur.fetchall()}
    return _SCHEMA_CACHE[table_name]


def _table_has_column(cur, table_name, column_name):
    return column_name in _get_table_columns(cur, table_name)


def _get_current_user_email():
    try:
        identity = get_jwt_identity()
        return identity if isinstance(identity, str) else identity.get('email')
    except Exception:
        return None


def _get_user_company_id(cur, user_email):
    if not user_email or not _table_has_column(cur, 'users', 'company_id'):
        return None
    cur.execute('SELECT company_id FROM users WHERE email = %s', (user_email,))
    row = cur.fetchone()
    return row[0] if row and row[0] is not None else None


# Optional modules gated by license (companies.enabled_modules JSONB) — e.g. 'log_de_patrullas'
# is not part of the standard offering and must be excluded from Reportes unless enabled.
OPTIONAL_MODULE_FORM_TYPES = {'log_de_patrullas'}


def _module_enabled(cur, company_id, module_key):
    if company_id is None:
        return False
    cur.execute('SELECT enabled_modules FROM companies WHERE id = %s', (company_id,))
    row = cur.fetchone()
    return bool(row and row[0] and module_key in row[0])


# --- Technical / System columns that should never be displayed in functional report details ---
TECHNICAL_SYSTEM_COLUMNS = {
    'company_id', 'customer_company_id', 'id_propiedad', 'cliente_instalacion',
    'latitude', 'latitud', 'longitude', 'longitud', 'location_accuracy', 'precision',
    'submitter_timezone', 'user_name', 'cliente_nombre', 'propiedad_nombre',
    'editado', 'editado_por', 'editado_en', 'motivo_edicion', 'motivo_detalle',
    'csrf_token', 'session_token', 'personas_involucradas', 'personas_data',
    'id', 'id_reporte_incidente', 'id_encuesta', 'id_supervision', 'id_informe',
    'id_patrulla', 'id_capacitacion', 'id_visita', 'id_planilla_vehicular'
}

# --- Form Configurations ---
FORM_CONFIGS = {
    'reporte_incidente': {
        'table': 'reportes_incidentes',
        'id_col': 'id_reporte_incidente',
        'date_col': 'creado_en',
        'user_col': 'user_email',
        'title_prefix': 'Reporte de Incidente',
        'joins': """
            LEFT JOIN users u ON t.user_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
            LEFT JOIN LATERAL (
                SELECT json_agg(json_build_object('tipo', rip.persona_tipo, 'nombre', rip.persona_nombre) ORDER BY rip.id) AS personas_data
                FROM reportes_incidentes_personas rip
                WHERE rip.id_reporte_incidente = t.id_reporte_incidente
            ) r_personas ON TRUE
        """,
        'columns': """
            t.creado_en,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            r_personas.personas_data AS personas_involucradas,
            u.name AS user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Puesto o Área Específica": "puesto_area_especifica",
            "Fecha y Hora": "fecha_hora",
            "Rol del Aplicador/Responsable": "rol_aplicador",
            "Turno": "turno",
            "Hora de Ingreso": "hora_entrada",
            "Hora de Salida": "hora_salida",
            "Nombre del Responsable": "nombre_responsable",
            "Número de Empleado": "numero_empleado",
            "Categoría": "categoria",
            "Razón de Ausentismo": "razon_ausentismo",
            "¿Se cubre el puesto?": "cubre_puesto",
            "Nombre persona que cubre": "nombre_persona_cubre",
            "Número de empleado (cubre)": "numero_empleado_cubre",
            # 2. Detalles del Incidente
            "Tipo de Incidente": "tipo_incidente",
            "Descripción": "descripcion_incidente",
            "Nivel de Severidad": "nivel_severidad",
            "Impacto": "impacto",
            "Descripción del Impacto": "descripcion_impacto",
            "Personas Involucradas": "personas_involucradas",
            "¿Reportado a autoridades?": "reportado_autoridades",
            "Número de reporte / radicado": "numero_reporte_autoridades",
            # 3. Plan de Acción y Seguimiento
            "Plan de Acción": "plan_accion",
            "Responsable Plan de Acción": "nombre_responsable_plan",
            "Fecha de Cumplimiento": "fecha_cumplimiento_plan",
            "Estado / Seguimiento": "estado_seguimiento_plan",
            # 4. Firmas y Evidencias
            "Firma": "firma_responsable",
            "Foto Evidencia": "foto_evidencia_url",
            "URLs de Imágenes o PDFs": "foto_evidencia_url"
        }
    },
    'medicion_experiencia_cliente': {
        'table': 'medicion_experiencia_cliente',
        'id_col': 'id_encuesta',
        'date_col': 'creado_en',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Medición Experiencia Cliente',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.creado_en,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name as user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Fecha y Hora": "fecha_hora",
            "Rol del Aplicador/Responsable": "rol_aplicador",
            "Nombre del Responsable": "nombre_responsable",
            # 2. Evaluación de la Seguridad (1-5)
            "Atención al Cliente": "atencion_cliente",
            "Comunicación": "comunicacion",
            "Confiabilidad": "confiabilidad",
            "Capacidad de Reacción": "capacidad_reaccion",
            "Cumplimiento": "cumplimiento",
            "Competencia del Personal": "competencia_personal",
            "Actitud de Servicio": "actitud_servicio",
            "Atención de Quejas": "atencion_quejas",
            # 3. Satisfacción Global
            "Calificación Global / NPS": "calificacion_global_nps",
            "¿Recomendaría el Servicio?": "recomendaria_servicio",
            # 4. Observaciones
            "Observaciones del Cliente": "observaciones_cliente",
            # 5. Validación y Firmas
            "Encuestado": "encuestado",
            "Firma Encuestado": "firma_encuestado",
            "Firma Responsable": "firma_responsable"
        }
    },
    'supervision_puesto': {
        'table': 'supervision_puesto',
        'id_col': 'id_supervision',
        'date_col': 'creado_en',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Control de Supervisión',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.creado_en,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name AS user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Fecha y Hora": "fecha_hora",
            "Nombre Supervisor": "supervisor",
            "Rol del Aplicador/Responsable": "rol_aplicador",
            # 2. Datos del Puesto
            "Puesto o Área Específica": "puesto_area_especifica",
            "Horario del Servicio": "horario_servicio",
            "Hora de Ingreso": "hora_entrada",
            "Hora de Salida": "hora_salida",
            "Tipo de Instalación": "tipo_servicio",
            "Modalidad de Servicio": "modalidad_servicio",
            # 3. Datos del Guardia de Seguridad
            "Nombre del Guardia": "nombre_guardia",
            "Número de Empleado": "numero_empleado",
            "Documento del Guardia": "documento_guardia",
            "Tiempo en el Puesto (meses)": "tiempo_en_puesto",
            "No. Licencia Portar Arma": "licencia_portar_arma",
            # 4. Inducción y Equipamiento
            "¿Se realiza inducción al puesto?": "realiza_induccion",
            "¿Conoce Órdenes y Consignas?": "conoce_ordenes_consignas",
            "¿Horario y detalles claros?": "horario_detalles_claros",
            "Porta arma en el puesto": "porta_arma",
            "Tipo de Arma": "tipo_arma",
            "Serie del Arma": "serie_arma",
            "Matrícula / Permiso": "matricula_arma",
            "Cantidad de Munición": "cantidad_municion",
            "Vencimiento Permiso de Porte": "fecha_vencimiento_permiso_porte",
            "Último Mtto (Arma)": "fecha_ultimo_mtto_arma",
            "Radio Asignado (Serial)": "radio_asignado_serial",
            "Marca del Radio": "marca_radio",
            "Tipo de Radio": "tipo_radio",
            "Último Mtto (Radio)": "fecha_ultimo_mtto_radio",
            # 5. Inspección de Puesto
            "Asistencia y Puntualidad": "asistencia_puntualidad",
            "Presentación y Uniforme": "presentacion_uniforme",
            "Prendas con Problemas": "problemas_uniforme",
            "Estado y Limpieza del Puesto": "estado_limpieza_puesto",
            "Equipamiento Completo": "equipamiento_completo",
            "Estado de Bitácora y Registros": "estado_bitacora",
            # 6. Observaciones, Evidencias y Firmas
            "Observaciones / Novedades": "observaciones_novedades",
            "Foto Evidencia": "foto_evidencia_url",
            "URLs de Imágenes o PDFs": "foto_evidencia_url",
            "Firma Supervisor": "firma_supervisor",
            "Firma del Guardia": "firma_guardia"
        }
    },
    'informe_novedades_disciplinario': {
        'table': 'informe_novedades_disciplinario',
        'id_col': 'id_informe',
        'date_col': 'creado_en',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Informe Disciplinario',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.creado_en,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name as user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Puesto o Área Específica": "puesto_area_especifica",
            "Fecha y Hora": "fecha_hora",
            "Rol del Aplicador/Responsable": "rol_aplicador",
            "Turno": "turno",
            "Hora de Ingreso": "hora_entrada",
            "Hora de Salida": "hora_salida",
            "Nombre del Responsable": "nombre_responsable",
            # 2. Datos del Empleado Involucrado
            "Nombre del Colaborador": "empleado_nombre",
            "Número de Empleado": "empleado_numero",
            "Documento del Colaborador": "empleado_documento",
            "Cargo del Colaborador": "empleado_cargo",
            # 3. Tipo de Novedad / Falta Disciplinaria
            "Tipo de Novedad / Falta": "tipo_novedad",
            # 4. Descripción de la Novedad
            "Sitio donde ocurrió": "sitio_ocurrencia",
            "Descripción de la Novedad": "descripcion_novedad",
            "Otras Personas Involucradas": "otras_personas_involucradas",
            # 5. Anexos / Evidencias
            "Anexos / Evidencias": "anexos",
            "Foto Evidencia": "anexos",
            # 6. Firmas
            "Firma del Responsable": "firma_responsable",
            "Firma del Colaborador": "firma_recibido_revisado",
            "¿Empleado se niega a firmar?": "empleado_niega_firmar",
            "Nombre del Testigo": "nombre_testigo",
            "Firma del Testigo": "firma_testigo"
        }
    },
    'log_de_patrullas': {
        'table': 'log_de_patrullas',
        'id_col': 'id_patrulla',
        'date_col': 'creado_en',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Log de Patrullas',
        'joins': "LEFT JOIN users u ON t.submitted_by_email = u.email",
        'columns': "t.creado_en, t.*, u.name as user_name",
        'data_mapping': {
            # 1. Datos de la Patrulla
            "ID Guardia / Nombre Guardia": "id_guardia_nombre_guardia",
            "Sitio / Ubicación": "sitio_ubicacion",
            "ID Patrulla": "id_patrulla_consecutivo",
            "Fecha": "fecha",
            "Hora Inicio": "hora_inicio",
            "Hora Fin": "hora_fin",
            # 2. Detalles de la Patrulla
            "Detalles Incidente": "detalles_incidente",
            "Riesgo Detectado": "riesgo_detectado",
            "Nivel de Riesgo": "nivel_riesgo",
            "Estado Patrulla": "estado_patrulla",
            "Contexto / Observaciones": "contexto_observaciones",
            # 3. Firmas
            "Firma Guardia": "firma_guardia",
            "Firma Supervisor": "firma_supervisor"
        }
    },
    'registro_de_capacitaciones': {
        'table': 'registro_de_capacitaciones',
        'id_col': 'id_capacitacion',
        'date_col': 'creado_en',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Registro de Capacitaciones',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.creado_en,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name as user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Fecha y Hora": "fecha_hora",
            "Tema de la Capacitación": "nombre_capacitacion",
            "Nombre del Responsable/Capacitador": "nombre_responsable",
            "Cargo del Responsable": "cargo_responsable",
            "Firma del Responsable": "firma_responsable",
            "Fotos/Documentos": "foto_evidencia_url",
            "URLs de Imágenes o PDFs": "foto_evidencia_url",
            # 2. Asistencia
            "Lista de Asistencia": "lista_asistencia"
        }
    },
    'registro_y_acta_de_visita': {
        'table': 'registro_y_acta_de_visita',
        'id_col': 'id_visita',
        'date_col': 'creado_en',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Acta de Visita',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.creado_en,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name as user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Fecha y Hora": "fecha_hora",
            "Motivo de la Visita": "motivo_visita",
            # 2. Agenda y Compromisos
            "Temas Tratados": "temas_tratados",
            "Acuerdos y Compromisos": "acuerdos_compromisos",
            "Nombre Responsable": "nombre_responsable",
            "Fecha de Cumplimiento": "fecha_cumplimiento",
            "Estado / Seguimiento": "compromisos_estados",
            # 3. Participantes y Firmas
            "Nombre del Visitante": "nombre_visitante",
            "Cargo del Visitante": "cargo_visitante",
            "Firma del Visitante": "firma_visitante",
            "Participantes del Cliente": "detalles_participantes"
        }
    },
    'planilla_vehicular': {
        'table': 'planilla_vehicular',
        'id_col': 'id_planilla_vehicular',
        'date_col': 'creado_en',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Planilla de Chequeo Pre-Operacional Vehicular',
        'sheet_title': 'Pre-Operacional Vehicular',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.creado_en,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name as user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Rol del Aplicador/Responsable": "rol_aplicador",
            "Nombre del Responsable": "nombre_responsable",
            "Fecha y Hora": "fecha_hora",
            "Turno": "turno",
            "Hora de Ingreso": "hora_entrada",
            "Hora de Salida": "hora_salida",
            "Número de Empleado": "numero_empleado",
            "Vehículo (Propio/Empresa)": "vehiculo_tipo",
            "Placa": "placa_vehiculo",
            "Fecha Último Mantenimiento": "fecha_ultimo_mantenimiento",
            "Último Kilometraje Registrado": "kilometraje_anterior",
            "Kilometraje Actual": "kilometraje_vehiculo",
            "Kilometraje Recorrido": "kilometraje_recorrido",
            # 2. Matriz de Inspección
            "Estado Rines": "estado_rines",
            "Juego Señales Carretera": "juego_senales_carretera",
            "Gato Hidráulico": "gato_hidraulico",
            "Palanca Gato": "palanca_gato",
            "Estado Asientos": "estado_asientos",
            "Estado Tapetes / Alfombras": "estado_tapetes_alfombras",
            "Limpieza Carrocería": "limpieza_carroceria",
            "Luces Delanteras": "luces_delanteras",
            "Luces Direccionales": "luces_direccionales",
            "Luces Traseras": "luces_traseras",
            "Parabrisas Delantero": "parabrisas_delantero",
            "Parabrisas Trasero": "parabrisas_trasero",
            "Defensa Delantera": "defensa_delantera",
            "Defensa Trasera": "defensa_trasera",
            "Puertas / Vidrios": "puertas_vidrios",
            "Tapa Radiador": "tapa_radiador",
            "Tapa Aceite Motor": "tapa_aceite_motor",
            "Batería Tapa": "bateria_tapa",
            "Espejo Retrovisor Interno": "espejo_retrovisor_interno",
            "Espejos Retrovisores Externos": "espejos_retrovisores_externos",
            "Limpiabrisas": "limpia_brisas",
            "Antena Radio": "antena_radio",
            "Radio Funciona": "radio_funciona",
            "Llanta Repuesto": "llanta_repuesto",
            "Aire Acondicionado": "aire_acondicionado",
            # 3. Evidencias, Diagrama y Novedades
            "Foto Frente": "foto_frente_url",
            "Foto Atrás": "foto_atras_url",
            "Foto Lado Derecho": "foto_lado_derecho_url",
            "Foto Lado Izquierdo": "foto_lado_izquierdo_url",
            "Diagrama de Daños": "diagrama_danos",
            "Novedades Críticas": "novedades_criticas",
            "Acción Inmediata": "accion_inmediata",
            # 4. Entrega y Cierre
            "Kilometraje de Entrega": "kilometraje_entrega",
            "Firma Entrega": "firma_entrega",
            "Kilometraje de Salida": "kilometraje_salida",
            "Firma Recibe": "firma_recibe",
            "Oficial de Operaciones": "oficial_operaciones_nombre",
            "Firma Oficial": "oficial_operaciones_firma",
            "Firma Responsable": "firma_responsable"
        }
    },
    'planilla_motocicletas': {
        'table': 'planilla_motocicletas',
        'id_col': 'id',
        'date_col': 'creado_en',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Planilla de Chequeo Pre-Operacional de Motocicletas',
        'sheet_title': 'Pre-Operacional Motocicletas',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.creado_en,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name as user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Rol del Aplicador/Responsable": "rol_aplicador",
            "Nombre del Responsable": "nombre_responsable",
            "Fecha y Hora": "fecha_hora",
            "Turno": "turno",
            "Hora de Ingreso": "hora_entrada",
            "Hora de Salida": "hora_salida",
            "Número de Empleado": "numero_empleado",
            "Placa": "placa_motocicleta",
            "Fecha Último Mantenimiento": "fecha_ultimo_mantenimiento",
            "Último Kilometraje Registrado": "kilometraje_anterior",
            "Kilometraje Actual": "kilometraje_motocicleta",
            "Kilometraje Recorrido": "kilometraje_recorrido",
            # 2. Matriz de Inspección
            "Estado Neumáticos": "estado_neumaticos",
            "Estado Rines": "estado_rines",
            "Equipo Carretera": "equipo_carretera",
            "Kit de Arrastre": "estado_kit_arrastre",
            "Palanca Soporte": "estado_palanca_soporte",
            "Forro Asiento": "estado_forro_asiento",
            "Tapas Derecha": "estado_tapas_derecha",
            "Luces Direccionales Derecha": "estado_luces_direccionales_derecha",
            "Luces Delanteras": "estado_luces_delanteras",
            "Guardafango Delantero": "estado_guarda_fango_delantero",
            "Sistema Freno Delantero": "estado_sistema_freno_delantero",
            "Manillar Embrague": "estado_manillar_embrague",
            "Manillar Freno Delantero": "estado_manillar_freno_delantero",
            "Manómetros / Indicadores": "estado_manometros_indicadores",
            "Tanque Combustible": "estado_tanque_combustible",
            "Tapa Tanque Combustible": "tapa_tanque_combustible",
            "Espejos Retrovisores": "espejos_retrovisores",
            "Tapa Aceite Motor": "tapa_aceite_motor",
            "Batería Tapa": "bateria_tapa",
            "Luces Izquierda": "estado_luces_izquierda",
            "Luces Direccionales Izquierda": "estado_luces_direccionales_izquierda",
            "Luz Trasera": "estado_luz_trasera",
            "Guardafango Trasero": "estado_guarda_fango_trasero",
            "Tubo de Escape": "estado_tubo_escape",
            "Palanca Freno": "estado_palanca_freno",
            "Palanca Cambios": "estado_palanca_cambios",
            # 3. Evidencias y Novedades
            "Foto Frente": "foto_frente_url",
            "Foto Atrás": "foto_atras_url",
            "Foto Lado Derecho": "foto_lado_derecho_url",
            "Foto Lado Izquierdo": "foto_lado_izquierdo_url",
            "Novedades Críticas": "novedades_criticas_detectadas",
            "Acción Inmediata": "accion_inmediata_tomada",
            # 4. Entrega y Cierre
            "Kilometraje de Entrega": "kilometraje_entrega",
            "Firma Entrega": "firma_entrega",
            "Kilometraje de Salida": "kilometraje_salida",
            "Firma Recibe": "firma_recibe",
            "Oficial de Operaciones": "oficial_operaciones_nombre",
            "Firma Oficial": "oficial_operaciones_firma",
            "Firma Responsable": "firma_responsable"
        }
    },
    'checklist_cumplimiento': {
        'table': 'checklist_cumplimiento',
        'id_col': 'id',
        'date_col': 'created_at',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Checklist Cumplimiento',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.created_at,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name as user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Puesto o Área Específica": "puesto_area_especifica",
            "Fecha y Hora": "fecha_hora",
            "Turno": "turno",
            "Rol del Aplicador/Responsable": "rol_aplicador",
            "Auditor": "nombre_auditor",
            # 2. Datos del Agente Supervisado
            "Agente Supervisado": "agente_nombre_completo",
            "Tipo de Documento": "agente_tipo_documento",
            "Número de Documento": "agente_numero_documento",
            "Cargo / Rol del Agente": "agente_cargo_rol",
            "Número de Empleado": "agente_numero_empleado",
            "Puesto del Agente": "agente_puesto",
            # 3. Documentación y Certificaciones
            "Curso / Certificación": "curso_certificacion",
            "Academia que Certifica": "academia_certifica",
            "Nro. Resolución": "nro_resolucion",
            "Fecha de Resolución": "fecha_resolucion",
            "Vigencia Desde": "vigencia_desde",
            "Vigencia Hasta": "vigencia_hasta",
            "Nivel de Cumplimiento": "nivel_cumplimiento",
            "Copia Física de Certificados": "copia_certificados_fisica",
            "Certificados en Sistema": "certificados_cargados_sistema",
            "Documentación Coincide con HV": "documentacion_coincide_hv",
            "Fechas Vigentes": "fechas_vigentes",
            # 4. Firmas y Evidencias
            "Firma Auditor": "firma_auditor",
            "Firma Guardia Supervisado": "firma_guarda_supervisado",
            "Foto Evidencia": "evidencia_url",
            "URLs de Imágenes o PDFs": "evidencia_url"
        }
    },
    'confiabilidad_equipos': {
        'table': 'confiabilidad_equipos',
        'id_col': 'id',
        'date_col': 'fecha',
        'user_col': 'submitted_by_email',
        'title_prefix': 'Confiabilidad Equipos',
        'joins': """
            LEFT JOIN users u ON t.submitted_by_email = u.email
            LEFT JOIN propiedades p ON t.id_propiedad = p.id_propiedad
            LEFT JOIN customer_companies cc ON t.customer_company_id = cc.id
            LEFT JOIN customer_companies cc2 ON p.customer_company_id = cc2.id
            LEFT JOIN LATERAL (
                SELECT pl.id_propiedad, pl.nombre, pl.customer_company_id
                FROM propiedades pl
                LEFT JOIN customer_companies ccl ON pl.customer_company_id = ccl.id
                WHERE t.id_propiedad IS NULL
                  AND LOWER(TRIM(pl.nombre)) = LOWER(TRIM(t.cliente_instalacion))
                ORDER BY
                    CASE WHEN ccl.company_id = t.company_id THEN 0 ELSE 1 END,
                    pl.id_propiedad
                LIMIT 1
            ) p_legacy ON TRUE
            LEFT JOIN customer_companies cc3
              ON p_legacy.customer_company_id = cc3.id
        """,
        'columns': """
            t.fecha,
            t.*,
            COALESCE(
                NULLIF(TRIM(cc.name), ''),
                NULLIF(TRIM(cc2.name), ''),
                NULLIF(TRIM(cc3.name), '')
            ) AS cliente_nombre,
            COALESCE(
                NULLIF(TRIM(p.nombre), ''),
                NULLIF(TRIM(p_legacy.nombre), ''),
                NULLIF(TRIM(t.cliente_instalacion), '')
            ) AS propiedad_nombre,
            u.name as user_name
        """,
        'data_mapping': {
            # 1. Datos Generales
            "Cliente / Empresa": "cliente_nombre",
            "Propiedad / Instalación": "propiedad_nombre",
            "Sitio / Ubicación": "sitio",
            "Fecha": "fecha",
            "Hora": "hora",
            # 2. Inventario de Equipos
            "Inventario": "inventario",
            # 3. Responsables y Firmas
            "Técnico de Mantenimiento": "tecnico_mantenimiento",
            "Supervisor de Seguridad": "supervisor_seguridad",
            "Firma Técnico": "firma_tecnico",
            "Firma Supervisor": "firma_supervisor"
        }
    }
}

def _ensure_json_serializable(val):
    """Convert JSONB values that psycopg2 may return as Python str(list/dict) into proper objects."""
    if isinstance(val, (dict, list)):
        return val
    if isinstance(val, str):
        try:
            return json.loads(val)
        except (ValueError, TypeError):
            pass
        try:
            return ast.literal_eval(val)
        except Exception:
            pass
    return val

# ── Fechas: un solo criterio para el detalle, el PDF, el Excel y el correo ──
# El PDF ya convertía a la hora de operación; el JSON que alimenta el detalle
# devolvía `str(datetime)`, o sea UTC cruda. El mismo instante salía como
# "2026-09-01 12:05:00+00:00" en pantalla y "01/09/2026 07:05 a. m." en el PDF.
# Compartir la regla es lo que impide que vuelvan a separarse.
_DATE_KEY_HINTS = ('fecha', 'vigencia')


def _es_clave_de_fecha(key):
    """True si la etiqueta de un campo corresponde a una fecha.

    Deliberadamente no incluye 'hora': "Horario del Servicio" vale 'Diurno' y
    no debe pasar por el formateador.
    """
    k = str(key or '').lower()
    return any(h in k for h in _DATE_KEY_HINTS)


def _localizar_fechas(data, tz):
    """Pasa a la hora local de la operación los campos de fecha ya serializados.

    Se aplica al construir el registro, así que el detalle, el PDF, el Excel y
    el correo parten del mismo texto. `format_local_datetime` es idempotente,
    de modo que las vistas que ya lo llamaban siguen funcionando igual.
    """
    from admin_bp import format_local_datetime
    for k, v in list(data.items()):
        if isinstance(v, str) and v and _es_clave_de_fecha(k):
            data[k] = format_local_datetime(v, tz=tz, time_sep=" ") or v
    return data


_EXPORT_BLANK_TEXT = {'', 'N/A', 'None', 'null', '[]', '{}'}

def _is_blank_export_value(value):
    """True when a field holds nothing the form actually captured.

    A record carries every column of its table, including the ones the current
    form no longer asks for, and the fetch helpers turn those NULLs into the
    "N/A" placeholder.  Exports use this to leave them out instead of printing
    rows and columns that the form being exported never had.

    Zero and False are answers, not blanks, so only text-ish emptiness counts.
    """
    if value is None:
        return True
    if isinstance(value, (bool, int, float, Decimal)):
        return False
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) == 0
    return str(value).strip() in _EXPORT_BLANK_TEXT

_INV_LABELS = {
    'tipo_equipo':           'Tipo de Equipo',
    'total_equipos':         'Total',
    'equipos_operativos':    'Operativos',
    'equipos_con_falla':     'Con Falla',
    'pendiente_reparacion':  'Pend. Reparación',
    'pendiente_compra':      'Pend. Compra',
    'comentario':            'Comentario',
}

def _format_structured_value_as_text(val):
    """Convert a structured list/dict into a clean, readable multiline text block (useful for Excel export)."""
    val = _ensure_json_serializable(val)
    if isinstance(val, list):
        items = []
        for i, item in enumerate(val, 1):
            if isinstance(item, dict):
                parts = []
                if 'tipo' in item and 'nombre' in item:
                    tipo_str = item.get('tipo') or 'Persona'
                    nombre_str = item.get('nombre') or ''
                    parts.append(f"{tipo_str}: {nombre_str}")
                else:
                    for k in ['tipo_equipo', 'total_equipos', 'equipos_operativos', 'equipos_con_falla', 'pendiente_reparacion', 'pendiente_compra', 'comentario']:
                        if k in item:
                            label = _INV_LABELS.get(k, k.replace('_', ' ').capitalize())
                            parts.append(f"{label}: {item[k]}")
                    # Add any other keys not in our preferred ordering
                    for k, v in item.items():
                        if k not in _INV_LABELS:
                            if 'firma' in k.lower() and isinstance(v, str) and v.startswith('data:image'):
                                v = '[Firma adjunta]'
                            parts.append(f"{k.replace('_', ' ').capitalize()}: {v}")
                items.append(f"[{i}] " + ", ".join(parts))
            else:
                items.append(str(item))
        return "\n".join(items)
    elif isinstance(val, dict):
        parts = []
        if 'tipo' in val and 'nombre' in val:
            return f"{val.get('tipo')}: {val.get('nombre')}"
        for k in ['tipo_equipo', 'total_equipos', 'equipos_operativos', 'equipos_con_falla', 'pendiente_reparacion', 'pendiente_compra', 'comentario']:
            if k in val:
                label = _INV_LABELS.get(k, k.replace('_', ' ').capitalize())
                parts.append(f"{label}: {val[k]}")
        for k, v in val.items():
            if k not in _INV_LABELS:
                if 'firma' in k.lower() and isinstance(v, str) and v.startswith('data:image'):
                    v = '[Firma adjunta]'
                parts.append(f"{k.replace('_', ' ').capitalize()}: {v}")
        return ", ".join(parts)
    return str(val)

def _render_inventario_html_table(value, is_email=False):
    """Generate HTML table for the Inventario field for WeasyPrint PDF or HTML Emails."""
    items = _ensure_json_serializable(value)
    if not isinstance(items, list) or not items:
        return ""
    
    headers = ['tipo_equipo', 'total_equipos', 'equipos_operativos', 'equipos_con_falla', 'pendiente_reparacion', 'pendiente_compra', 'comentario']
    
    # Determine active headers
    all_keys = set()
    for item in items:
        if isinstance(item, dict):
            all_keys.update(item.keys())
            
    active_headers = []
    for h in headers:
        if h in all_keys:
            active_headers.append(h)
    for k in all_keys:
        if k not in headers:
            active_headers.append(k)
            
    if not active_headers:
        return ""
        
    # Styles matching standard premium look
    if is_email:
        th_style = "border: 1px solid #cbd5e1; padding: 6px 8px; text-align: left; background-color: #f1f5f9; font-weight: bold; color: #374151; font-size: 11px;"
        td_style = "border: 1px solid #cbd5e1; padding: 6px 8px; text-align: left; color: #1f2937; font-size: 11px; vertical-align: top;"
        table_style = "width: 100%; border-collapse: collapse; margin-top: 8px; background-color: #ffffff; font-family: Arial, Helvetica, sans-serif;"
    else:
        th_style = "border: 1px solid #cbd5e1; padding: 4px 6px; text-align: left; background-color: #f1f5f9; font-weight: bold; color: #374151; font-size: 7.5pt;"
        td_style = "border: 1px solid #cbd5e1; padding: 4px 6px; text-align: left; color: #1f2937; font-size: 7.5pt; vertical-align: top;"
        table_style = "width: 100%; border-collapse: collapse; margin-top: 5px; background-color: #ffffff; font-family: Arial, Helvetica, sans-serif;"
    
    # Build header
    table_hdr_cells = []
    for h in active_headers:
        lbl = _INV_LABELS.get(h, h.replace('_', ' ').capitalize())
        table_hdr_cells.append(f'<th style="{th_style}">{lbl}</th>')
    table_hdr_html = "".join(table_hdr_cells)
    
    # Build body
    table_body_rows = []
    for item in items:
        if not isinstance(item, dict):
            continue
        row_cells = []
        for h in active_headers:
            v = item.get(h, '')
            v_str = str(v).strip() if v is not None else ''
            if v_str == '':
                v_str = '—'
            row_cells.append(f'<td style="{td_style}">{v_str}</td>')
        table_body_rows.append(f'<tr>{"".join(row_cells)}</tr>')
    table_body_html = "".join(table_body_rows)
    
    return f'<table style="{table_style}"><thead><tr>{table_hdr_html}</tr></thead><tbody>{table_body_html}</tbody></table>'

def fetch_reports(offset, limit, filters=None, form_type='all', skip_signing=False):
    conn = None
    cur = None
    all_reports = []
    
    # Determine which form types to fetch
    types_to_fetch = []
    if not form_type or form_type == 'all' or form_type == '':
        types_to_fetch = list(FORM_CONFIGS.keys())
    elif form_type in FORM_CONFIGS:
        types_to_fetch = [form_type]
    else:
        app_logger.warning(f"Invalid form_type: {form_type}. Defaulting to all.")
        types_to_fetch = list(FORM_CONFIGS.keys())

    conn = get_db_connection()
    if not conn:
        app_logger.error("Failed to get database connection in fetch_reports. Returning empty list.")
        return [], 0

    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        user_email = _get_current_user_email()
        company_id = _get_user_company_id(cur, user_email)

        # Exclude optional modules (e.g. log_de_patrullas) not enabled for this company's license.
        types_to_fetch = [
            t for t in types_to_fetch
            if t not in OPTIONAL_MODULE_FORM_TYPES or _module_enabled(cur, company_id, t)
        ]

        total_count = 0

        for f_type in types_to_fetch:
            config = FORM_CONFIGS[f_type]
            
            # Build WHERE clause based on filters
            where_conditions = []
            query_params = []

            # Tenant isolation — always applied regardless of other filters
            if company_id is not None:
                if not _table_has_column(cur, config['table'], 'company_id'):
                    current_app.logger.error(
                        "Security: table '%s' missing company_id — refusing cross-tenant query",
                        config['table']
                    )
                    return [], 0
                # Legacy rows can have no company_id.  Attribute them through the
                # submitter already joined as ``u`` instead of hiding them or
                # admitting every unscoped row.
                where_conditions.append(
                    "(t.company_id = %s OR (t.company_id IS NULL AND u.company_id = %s))"
                )
                query_params.extend([company_id, company_id])

            if filters:
                # Report ID filter
                if filters.get('report_id'):
                    try:
                        report_id = int(filters['report_id'])
                        where_conditions.append(f"t.{config['id_col']} = %s")
                        query_params.append(report_id)
                    except (ValueError, TypeError):
                        pass # Ignore invalid ID for this type

                # Submitted by filter
                if filters.get('submitted_by'):
                    where_conditions.append(f"(LOWER(u.name) LIKE LOWER(%s) OR LOWER(t.{config['user_col']}) LIKE LOWER(%s))")
                    search_term = f"%{filters['submitted_by']}%"
                    query_params.extend([search_term, search_term])

                # Date range filters
                if filters.get('start_date'):
                    where_conditions.append(f"t.{config['date_col']} >= %s")
                    query_params.append(filters['start_date'])

                if filters.get('end_date'):
                    where_conditions.append(f"t.{config['date_col']} <= %s")
                    query_params.append(filters['end_date'])

                # Cliente / Empresa y Propiedad / Instalación, para TODOS los tipos.
                # Antes esto estaba dentro de `if f_type == 'reporte_incidente'`, de
                # modo que filtrar por propiedad no acotaba ningún otro formulario.
                # Se reutiliza el mismo helper que los dashboards, así que un
                # registro legacy sin FK se resuelve igual aquí que allá.
                if filters.get('cliente') or filters.get('property_id') or filters.get('property'):
                    from dashboard_bp import _add_scope_filters
                    # log_de_patrullas guarda el sitio en `sitio`, no en
                    # cliente_instalacion: sin esto la consulta apuntaría a una
                    # columna inexistente y ese tipo de reporte reventaría.
                    col_inst = None if config['table'] == 'log_de_patrullas' else 'cliente_instalacion'
                    _add_scope_filters(
                        where_conditions, query_params,
                        cliente=filters.get('cliente'),
                        propiedad=filters.get('property_id') or filters.get('property'),
                        col_cliente_inst=col_inst,
                        col_puesto=None,
                        prefix='t.',
                    )

                # Lugar del incidente: solo reporte_incidente tiene esa columna.
                if f_type == 'reporte_incidente' and filters.get('location'):
                    where_conditions.append("LOWER(t.puesto_area_especifica) LIKE LOWER(%s)")
                    query_params.append(f"%{filters['location']}%")

            where_clause = ""
            if where_conditions:
                where_clause = "WHERE " + " AND ".join(where_conditions)

            # Count Query
            count_query = f"SELECT COUNT(*) FROM {config['table']} t {config['joins']} {where_clause}"
            cur.execute(count_query, tuple(query_params))
            total_count += cur.fetchone()[0]

            # Data Query
            # Fetch enough rows to satisfy the global offset + limit
            fetch_limit = limit + offset
            query = f"""
                SELECT {config['columns']}
                FROM {config['table']} t
                {config['joins']}
                {where_clause}
                ORDER BY t.{config['date_col']} DESC NULLS LAST, t.{config['id_col']} DESC
                LIMIT %s
            """
            
            cur.execute(query, tuple(query_params + [fetch_limit]))
            rows = cur.fetchall()

            # Una sola vez por consulta: get_operation_timezone lee los umbrales
            # y con ellos toca la base, así que no puede ir dentro del bucle.
            from admin_bp import get_operation_timezone, format_local_datetime as _fmt_local
            _tz_op = get_operation_timezone()

            for row_dict in rows:
                # Determine display name
                display_name = row_dict.get("user_name") or row_dict.get(config['user_col'], "desconocido")
                
                # Determine date
                date_val = row_dict.get(config['date_col'])
                if isinstance(date_val, datetime):
                    date_str = date_val.isoformat()
                else:
                    date_str = str(date_val) if date_val else "N/A"

                submitter_tz = row_dict.get('submitter_timezone') or 'UTC'

                # Map data
                mapped_data = {}
                processed_cols = set() # Keep track of columns already processed

                # 1. Process explicit data_mapping first
                for label, col_name in config['data_mapping'].items():
                    val = row_dict.get(col_name)
                    processed_cols.add(col_name) # Mark as processed

                    # Handle GCS URLs signing for specific labels
                    if label == 'URLs de Imágenes o PDFs' and val and not skip_signing:
                        signed_urls = []
                        for url in str(val).split('\n'):
                            url = url.strip()
                            if url:
                                signed_urls.append(generate_signed_url(url))
                        val = '\n'.join(signed_urls)
                    # Sign image and signature columns if they are GCS URLs
                    elif (_es_columna_de_imagen(col_name) and val and not skip_signing
                          and isinstance(val, str) and 'storage.googleapis.com' in val):
                        urls = str(val).split('\n')
                        signed_urls = []
                        for url in urls:
                            url_t = url.strip()
                            if 'storage.googleapis.com' in url_t:
                                signed_urls.append(generate_signed_url(url_t))
                            else:
                                signed_urls.append(url_t)
                        val = '\n'.join(signed_urls)

                    # Ensure JSON serializable
                    if isinstance(val, (datetime, date, time)):
                        val = str(val)
                    elif isinstance(val, Decimal):
                        val = float(val)
                    else:
                        val = _ensure_json_serializable(val)

                    mapped_data[label] = val

                # 2. Add unmapped fields, filtering out system and technical columns
                system_cols = {
                    config['id_col'], config['date_col'], config['user_col'], 'user_name', 'submitter_timezone',
                    'editado', 'editado_por', 'editado_en', 'motivo_edicion', 'motivo_detalle',
                }.union(TECHNICAL_SYSTEM_COLUMNS)
                for col_name, val in row_dict.items():
                    if col_name not in processed_cols and col_name.lower() not in system_cols:
                        # Sign signatures if they are GCS URLs
                        if (_es_columna_de_imagen(col_name) and val and not skip_signing
                                and isinstance(val, str) and 'storage.googleapis.com' in val):
                             val = generate_signed_url(val)

                        # Convert snake_case to Title Case for display
                        display_label = ' '.join(word.capitalize() for word in col_name.split('_'))

                        # Handle Date/Time objects for JSON serialization
                        if isinstance(val, (datetime, date, time)):
                            val = str(val)
                        else:
                            val = _ensure_json_serializable(val)

                        mapped_data[display_label] = val

                editado_en_val = row_dict.get('editado_en')
                report = {
                    "id": row_dict.get(config['id_col']),
                    "title": f"{config['title_prefix']} #{row_dict.get(config['id_col'])}",
                    "submittedBy": display_name,
                    "dateSubmitted": date_str,
                    "dateSubmittedLocal": _fmt_local(date_str, tz=_tz_op, time_sep=" "),
                    "submitterTimezone": submitter_tz,
                    "data": _localizar_fechas(mapped_data, _tz_op),
                    "formType": f_type,
                    "editado": bool(row_dict.get('editado')),
                    "editado_por": row_dict.get('editado_por') or '',
                    "editado_en": editado_en_val.strftime('%Y-%m-%d %H:%M') if hasattr(editado_en_val, 'strftime') else '',
                }
                all_reports.append(report)

        # Sort all reports by date descending
        all_reports.sort(key=lambda x: x['dateSubmitted'], reverse=True)
        
        # Apply pagination to the combined list
        start = offset
        end = offset + limit
        paginated_reports = all_reports[start:end]
        
        return paginated_reports, total_count

    except Exception as e:
        app_logger.error(f"Error fetching reports: {e}", exc_info=True)
        return [], 0
    finally:
        if conn:
            conn.close()

def fetch_reports_by_ids(report_ids, form_type='reporte_incidente', skip_signing=False):
    conn = None
    cur = None
    reports = []
    if not report_ids:
        app_logger.info("fetch_reports_by_ids received an empty report_ids list.")
        return []
        
    if form_type not in FORM_CONFIGS:
        app_logger.warning(f"Invalid form_type: {form_type}. Defaulting to reporte_incidente.")
        form_type = 'reporte_incidente'
        
    config = FORM_CONFIGS[form_type]

    try:
        conn = get_db_connection()
        if not conn:
            app_logger.error("Failed to get database connection in fetch_reports_by_ids. Returning empty list.")
            return reports

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        user_email = _get_current_user_email()
        company_id = _get_user_company_id(cur, user_email)

        if form_type in OPTIONAL_MODULE_FORM_TYPES and not _module_enabled(cur, company_id, form_type):
            app_logger.warning(f"fetch_reports_by_ids: module '{form_type}' not enabled for company {company_id}")
            return reports

        # Log the raw report_ids before cleaning
        app_logger.info(f"Raw report_ids received by fetch_reports_by_ids: {report_ids}")

        clean_ids = [int(id) for id in report_ids if isinstance(id, (int, str)) and str(id).isdigit()]
        
        # Log the cleaned report_ids
        app_logger.info(f"Cleaned report_ids for fetch_reports_by_ids: {clean_ids}")

        if not clean_ids:
            app_logger.warning("After cleaning, report_ids list is empty. No reports to fetch.")
            return []

        placeholders = ','.join(['%s'] * len(clean_ids))
        params = list(clean_ids)

        tenant_clause = ""
        if company_id is not None:
            if not _table_has_column(cur, config['table'], 'company_id'):
                current_app.logger.error(
                    "Security: table '%s' missing company_id — refusing cross-tenant query",
                    config['table']
                )
                return []
            tenant_clause = (
                "AND (t.company_id = %s OR "
                "(t.company_id IS NULL AND u.company_id = %s))"
            )
            params.extend([company_id, company_id])

        # ORDER BY: Sort by creation date (newest reports first - most recent submissions at the top)
        query = f"""
            SELECT {config['columns']}
            FROM {config['table']} t
            {config['joins']}
            WHERE t.{config['id_col']} IN ({placeholders}) {tenant_clause}
            ORDER BY t.{config['date_col']} DESC NULLS LAST, t.{config['id_col']} DESC
        """
        app_logger.info(f"Executing fetch_reports_by_ids query for IDs: {clean_ids} with form_type: {form_type}.")
        cur.execute(query, params)
        rows = cur.fetchall()
        app_logger.info(f"Fetched {len(rows)} specific reports.")
        
        from admin_bp import get_operation_timezone, format_local_datetime as _fmt_local
        _tz_op = get_operation_timezone()

        for row_dict in rows:
            # Determine display name
            display_name = row_dict.get("user_name") or row_dict.get(config['user_col'], "desconocido")
            
            # Determine date
            date_val = row_dict.get(config['date_col'])
            if isinstance(date_val, datetime):
                date_str = date_val.isoformat()
            else:
                date_str = str(date_val) if date_val else "N/A"

            submitter_tz = row_dict.get('submitter_timezone') or 'UTC'

            # Map data fields
            data_content = {}
            processed_cols = set()
            
            # 1. Process explicit data_mapping first
            for label, col_name in config['data_mapping'].items():
                val = row_dict.get(col_name)
                processed_cols.add(col_name)
                
                # Generate signed URLs for image/pdf columns
                if val and (label == "URLs de Imágenes o PDFs" or col_name in ('foto_evidencia_url', 'anexos')):
                    if not skip_signing:
                        urls = str(val).split('\n')
                        signed_urls = []
                        for url in urls:
                            signed_urls.append(generate_signed_url(url.strip()))
                        val = '\n'.join(signed_urls)
                # Sign image and signature columns if they are GCS URLs
                elif _es_columna_de_imagen(col_name) and val and isinstance(val, str) and 'storage.googleapis.com' in val:
                    if not skip_signing:
                        urls = str(val).split('\n')
                        signed_urls = []
                        for url in urls:
                            url_t = url.strip()
                            if 'storage.googleapis.com' in url_t:
                                signed_urls.append(generate_signed_url(url_t))
                            else:
                                signed_urls.append(url_t)
                        val = '\n'.join(signed_urls)

                if isinstance(val, (datetime, date, time)):
                    val = str(val)
                elif isinstance(val, Decimal):
                    val = float(val)
                elif val is not None:
                    val = _ensure_json_serializable(val)
                else:
                    val = "N/A"
                data_content[label] = val

            # 2. Add unmapped fields, filtering out system and technical columns
            system_cols = {
                config['id_col'], config['date_col'], config['user_col'], 'user_name', 'submitter_timezone',
                'editado', 'editado_por', 'editado_en', 'motivo_edicion', 'motivo_detalle',
            }.union(TECHNICAL_SYSTEM_COLUMNS)
            for col_name, val in row_dict.items():
                if col_name not in processed_cols and col_name.lower() not in system_cols:
                    # Sign signatures if they are GCS URLs
                    if _es_columna_de_imagen(col_name) and val and isinstance(val, str) and 'storage.googleapis.com' in val:
                         if not skip_signing:
                            val = generate_signed_url(val)

                    # Convert snake_case to Title Case for display
                    display_label = ' '.join(word.capitalize() for word in col_name.split('_'))

                    # Handle Date/Time objects for JSON serialization
                    if isinstance(val, (datetime, date, time)):
                        val = str(val)
                    elif isinstance(val, Decimal):
                        val = float(val)
                    elif val is not None:
                        val = _ensure_json_serializable(val)
                    else:
                        val = "N/A"

                    data_content[display_label] = val

            editado_en_val = row_dict.get('editado_en')
            forms_data = {
                "id": row_dict[config['id_col']],
                "title": f"{config['title_prefix']} #{row_dict[config['id_col']]}",
                "submittedBy": display_name,
                # ISO: se ordena por este campo como texto, así que no se
                # formatea. La versión legible viaja aparte.
                "dateSubmitted": date_str,
                "dateSubmittedLocal": _fmt_local(date_str, tz=_tz_op, time_sep=" "),
                "submitterTimezone": submitter_tz,
                "data": _localizar_fechas(data_content, _tz_op),
                "formType": form_type,
                "editado": bool(row_dict.get('editado')),
                "editado_por": row_dict.get('editado_por') or '',
                "editado_en": editado_en_val.strftime('%Y-%m-%d %H:%M') if hasattr(editado_en_val, 'strftime') else '',
            }
            reports.append(forms_data)

    except psycopg2.Error as e:
        app_logger.error(f"PostgreSQL Error in fetch_reports_by_ids: {e}", exc_info=True)
        reports = []
    except Exception as e:
        app_logger.error(f"An unexpected error occurred in fetch_reports_by_ids: {e}", exc_info=True)
        reports = []
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()
            app_logger.info("Database connection closed in fetch_reports_by_ids.")
    return reports

def fetch_reports_mixed(items, skip_signing=False):
    """
    Fetches reports from multiple tables based on a list of item dicts.
    Each item in 'items' should be a dict with: {'id': int/str, 'formType': str}
    Returns a unified list of report objects.
    """
    if not items:
        return []

    # Group IDs by formType
    ids_by_type = {}
    for item in items:
        f_type = item.get('formType', 'reporte_incidente')
        r_id = item.get('id')
        
        if not r_id:
            continue
            
        if f_type not in ids_by_type:
            ids_by_type[f_type] = []
        ids_by_type[f_type].append(r_id)

    all_reports = []
    
    # Fetch for each type
    for f_type, ids in ids_by_type.items():
        if not ids:
            continue
        app_logger.info(f"fetch_reports_mixed: Fetching {len(ids)} reports of type {f_type}")
        type_reports = fetch_reports_by_ids(ids, form_type=f_type, skip_signing=skip_signing)
        all_reports.extend(type_reports)

    # Sort combined results by date (submitted at)
    # The 'dateSubmitted' field is a string, so sorting might be imperfect if format varies, but usually YYYY-MM-DD HH:MM:SS
    all_reports.sort(key=lambda x: str(x.get('dateSubmitted', '')), reverse=True)
    
    return all_reports

def _get_email_password():
    """Fetch SMTP password from env or Secret Manager."""
    pw = os.environ.get('EMAIL_PASSWORD')
    if pw:
        return pw
    try:
        project_id = current_app.config.get('GCP_PROJECT_ID')
        if not project_id:
            return None
        from google.cloud import secretmanager
        client = secretmanager.SecretManagerServiceClient()
        name = f"projects/{project_id}/secrets/admin-email-pass/versions/latest"
        response = client.access_secret_version(request={"name": name})
        return response.payload.data.decode("UTF-8")
    except Exception as e:
        app_logger.error(f"Could not retrieve email password: {e}", exc_info=True)
        return None


def send_reports_email(recipient_email, subject, body, is_html=False):
    _email_username = current_app.config.get('EMAIL_USERNAME')
    _smtp_server = current_app.config.get('SMTP_SERVER')
    _smtp_port = int(current_app.config.get('SMTP_PORT', 587))
    _email_password = _get_email_password()

    if not all([_email_username, _email_password, _smtp_server, _smtp_port]):
        app_logger.error("Email sending skipped: Missing one or more email credentials or invalid port.")
        return False, "Missing or invalid email configuration."

    msg = MIMEMultipart()
    msg['From'] = _email_username
    msg['To'] = recipient_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html' if is_html else 'plain'))

    app_logger.info(f"Attempting to send email to {recipient_email} via {_smtp_server}:{_smtp_port} from {_email_username}.")
    try:
        server = None
        context = ssl.create_default_context()

        if _smtp_port == 465: # SMTP_SSL for port 465
            server = smtplib.SMTP_SSL(_smtp_server, _smtp_port, context=context, timeout=10)
        else: # Standard SMTP with STARTTLS for other ports like 587
            server = smtplib.SMTP(_smtp_server, _smtp_port, timeout=10)
            server.starttls(context=context)

        server.login(_email_username, _email_password)
        server.send_message(msg)
        server.quit()
        app_logger.info(f"Email sent successfully to {recipient_email}.")
        return True, "Email sent successfully."

    except smtplib.SMTPAuthenticationError:
        app_logger.error(f"SMTP Authentication Error: Check email username and password for {_email_username}.", exc_info=True)
        return False, "Authentication failed. Check email credentials."
    except smtplib.SMTPServerDisconnected:
        app_logger.error(f"SMTP Server Disconnected: Server {_smtp_server}:{_smtp_port} disconnected unexpectedly.", exc_info=True)
        return False, "The email server is unavailable. Please try again later."
    except ConnectionRefusedError:
        app_logger.error(f"SMTP Connection Refused: Check SMTP_HOST, SMTP_PORT, and firewall rules for {_smtp_server}:{_smtp_port}.", exc_info=True)
        return False, "Connection refused by the email server."
    except TimeoutError:
        app_logger.error(f"SMTP Connection Timeout: Could not connect to {_smtp_server}:{_smtp_port}. Check network connectivity and firewall rules.", exc_info=True)
        return False, "Connection timed out with the email server."
    except Exception as e:
        app_logger.error(f"An unexpected error occurred while sending email to {recipient_email}: {e}", exc_info=True)
        return False, f"An error occurred while sending email: {e}"

def admin_required(f):
    """
    Decorator that requires the user to be an admin, verified against the DB.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        is_api = request.path.startswith('/api/') or bool(
            request.accept_mimetypes and request.accept_mimetypes.accept_json
        )
        try:
            user_email = get_jwt_identity()
            conn = get_db_connection()
            if not conn:
                if is_api:
                    return jsonify({"error": "Service unavailable"}), 503
                return redirect('/')
            try:
                cur = conn.cursor()
                cur.execute('SELECT is_admin, is_active FROM users WHERE email = %s', (user_email,))
                row = cur.fetchone()
                cur.close()
            finally:
                conn.close()
            if not row or not row[0] or not row[1]:
                app_logger.warning("Admin access denied for %s", user_email)
                if is_api:
                    return jsonify({"error": "Acceso denegado"}), 403
                return redirect('/landing/')
            return f(*args, **kwargs)
        except Exception as e:
            app_logger.error("Error in admin_required decorator: %s", e, exc_info=True)
            if is_api:
                return jsonify({"error": "Authentication error"}), 500
            return redirect('/')

    return decorated_function

# --- Routes ---
@viewer_bp.route('/')
@jwt_required()
@admin_required
def index():
    user_email = get_jwt_identity()
    
    # Get JWT claims and admin status
    try:
        claims = get_jwt()
        user_name = claims.get('name', user_email.split('@')[0])
        is_admin = claims.get('is_admin', False)
        app_logger.info(f"Admin user {user_email} (is_admin={is_admin}) accessing viewer dashboard")
    except Exception as e:
        app_logger.error(f"Error getting JWT claims: {e}", exc_info=True)
        user_name = user_email.split('@')[0]
        is_admin = False
    
    # Get user name from database as fallback
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        if conn:
            cur = conn.cursor()
            cur.execute('SELECT "name" FROM "users" WHERE email = %s', (user_email,))
            user_row = cur.fetchone()
            if user_row and user_row[0]:
                user_name = user_row[0]
                app_logger.info(f"User found in DB: {user_name}")
    except Exception as e:
        app_logger.error(f"Error fetching user name: {e}", exc_info=True)
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

    initial_reports, total_reports_count = fetch_reports(offset=0, limit=10)
    return render_template("index.html", 
                         current_user=user_email, 
                         forms=initial_reports, 
                         user_name=user_name,
                         is_admin=is_admin,  # Pass admin status to template
                         total_reports=total_reports_count,
                         login_service_url='/',
                         landing_service_url='/landing/',
                         forms_service_url='/forms',
                         dashboard_service_url='/dashboard')

@viewer_bp.route('/api/properties', methods=['GET'])
@jwt_required()
def get_properties():
    """Get all properties from the database"""
    conn = None
    cur = None
    properties = []
    try:
        conn = get_db_connection()
        if not conn:
            app_logger.error("Failed to get database connection in get_properties.")
            return jsonify({"error": "Database connection failed", "properties": []}), 500

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        user_email = _get_current_user_email()
        company_id = _get_user_company_id(cur, user_email)
        
        # Mismo criterio que /forms/api/properties y /dashboard/api/properties:
        # solo instalaciones activas. Era el único listado del repo sin ese
        # filtro, y por eso Reportes seguía ofreciendo las propiedades dadas de
        # baja —los nombres viejos que ya nadie usa— junto a las vigentes.
        if company_id is not None and _table_has_column(cur, 'propiedades', 'customer_company_id'):
            query = """
                SELECT DISTINCT p.id_propiedad, p.nombre, p.direccion,
                       p.customer_company_id, cc.name AS cliente
                FROM propiedades p
                LEFT JOIN customer_companies cc ON cc.id = p.customer_company_id
                WHERE p.nombre IS NOT NULL
                  AND COALESCE(p.activa, TRUE) = TRUE
                  -- Sin el OR, la condición sobre cc convierte el LEFT JOIN en
                  -- INNER y desaparecen las instalaciones aún sin cliente.
                  AND (cc.company_id = %s OR p.customer_company_id IS NULL)
                ORDER BY p.nombre
            """
            cur.execute(query, (company_id,))
        else:
            query = """
                SELECT DISTINCT p.id_propiedad, p.nombre, p.direccion,
                       p.customer_company_id, cc.name AS cliente
                FROM propiedades p
                LEFT JOIN customer_companies cc ON cc.id = p.customer_company_id
                WHERE p.nombre IS NOT NULL
                  AND COALESCE(p.activa, TRUE) = TRUE
                ORDER BY p.nombre
            """
            cur.execute(query)
        
        app_logger.info("Fetching properties from database")
        rows = cur.fetchall()
        
        clientes = {}
        for row in rows:
            cid = row["customer_company_id"]
            cnombre = (row["cliente"] or '').strip()
            properties.append({
                "id": row["id_propiedad"],
                "name": row["nombre"],
                # Con qué encadenar Cliente / Empresa → Propiedad / Instalación.
                "cliente_id": cid,
                "cliente": cnombre,
                # Solo se muestra cuando dos instalaciones del mismo cliente
                # comparten nombre: es lo único que las distingue en pantalla.
                "direccion": (row["direccion"] or '').strip(),
            })
            if cid is not None and cnombre:
                clientes[cid] = cnombre

        app_logger.info(f"Retrieved {len(properties)} properties, {len(clientes)} clientes")
        return jsonify({
            "properties": properties,
            "clientes": [{"id": k, "name": v} for k, v in
                         sorted(clientes.items(), key=lambda kv: kv[1].lower())],
        }), 200
        
    except psycopg2.Error as e:
        app_logger.error(f"PostgreSQL Error in get_properties: {e}", exc_info=True)
        return jsonify({"error": "Database error", "properties": []}), 500
    except Exception as e:
        app_logger.error(f"Unexpected error in get_properties: {e}", exc_info=True)
        return jsonify({"error": "Server error", "properties": []}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

@viewer_bp.route('/api/locations', methods=['GET'])
@jwt_required()
def get_locations():
    """Get locations, optionally filtered by property"""
    property_id = request.args.get('property_id', type=int)
    
    conn = None
    cur = None
    locations = []
    try:
        conn = get_db_connection()
        if not conn:
            app_logger.error("Failed to get database connection in get_locations.")
            return jsonify({"error": "Database connection failed", "locations": []}), 500

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        user_email = get_jwt_identity()
        company_id = _get_user_company_id(cur, user_email)

        conds = [
            "puesto_area_especifica IS NOT NULL",
            "TRIM(puesto_area_especifica) <> ''",
        ]
        params = []
        if property_id:
            conds.append("id_propiedad = %s")
            params.append(property_id)
        if company_id is not None:
            conds.append("company_id = %s")
            params.append(company_id)

        where_clause = " AND ".join(conds)
        query = f"""
            SELECT DISTINCT TRIM(puesto_area_especifica) AS nombre
            FROM reportes_incidentes
            WHERE {where_clause}
            ORDER BY nombre
        """
        cur.execute(query, params)

        rows = cur.fetchall()

        for idx, row in enumerate(rows, start=1):
            locations.append({
                "id": idx,
                "name": row["nombre"]
            })

        app_logger.info("Retrieved %d locations for property_id: %s", len(locations), property_id)
        return jsonify({"locations": locations}), 200

    except psycopg2.Error as e:
        app_logger.error("PostgreSQL Error in get_locations: %s", e, exc_info=True)
        return jsonify({"error": "Database error", "locations": []}), 500
    except Exception as e:
        app_logger.error("Unexpected error in get_locations: %s", e, exc_info=True)
        return jsonify({"error": "Server error", "locations": []}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

@viewer_bp.route('/api/reports', methods=['GET'])
@jwt_required()
def get_more_reports():
    offset = request.args.get('offset', type=int, default=0)
    limit = request.args.get('limit', type=int, default=10)
    ids_only = request.args.get('ids_only', type=str, default='false').lower() == 'true'

    if offset < 0:
        offset = 0
    
    # Allow unlimited limit when ids_only is true, otherwise cap at 100
    if not ids_only:
        if limit <= 0 or limit > 100:
            limit = 10

    # Extract filter parameters
    filters = {}
    
    # Report ID filter
    if request.args.get('report_id'):
        filters['report_id'] = request.args.get('report_id')

    # Cliente / Empresa — acota antes que la propiedad y manda sobre ella.
    if request.args.get('cliente') or request.args.get('customer_company_id'):
        filters['cliente'] = request.args.get('cliente') or request.args.get('customer_company_id')

    # Property filters (ID takes precedence over name)
    if request.args.get('property_id'):
        filters['property_id'] = request.args.get('property_id')
    elif request.args.get('property'):
        filters['property'] = request.args.get('property')
    
    # Location filters (ID takes precedence over name)
    if request.args.get('location_id'):
        filters['location_id'] = request.args.get('location_id')
    elif request.args.get('location'):
        filters['location'] = request.args.get('location')
    
    # Other filters
    if request.args.get('submitted_by'):
        filters['submitted_by'] = request.args.get('submitted_by')
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')

    # Form Type Filter
    form_type = request.args.get('form_type', 'all')

    app_logger.info(f"API request with filters: {filters}, form_type: {form_type}, ids_only: {ids_only}")
    
    try:
        reports, total_count = fetch_reports(offset, limit, filters if filters else None, form_type=form_type)
        
        # If ids_only is true, return simplified response with just IDs and form types
        if ids_only:
            report_ids = [{'id': r['id'], 'formType': r['formType']} for r in reports]
            return jsonify({"report_ids": report_ids, "total_count": total_count})
        
        return jsonify({"reports": reports, "total_count": total_count})
    except psycopg2.Error as e:
        app_logger.error(f"Database error in get_more_reports: {e}", exc_info=True)
        return jsonify({"error": "Database error", "reports": [], "total_count": 0}), 500
    except Exception as e:
        app_logger.error(f"Server error in get_more_reports: {e}", exc_info=True)
        return jsonify({"error": "Server error", "reports": [], "total_count": 0}), 500

# New route to handle fetching a single report by ID, assumed to be used by "Ver Detalles"
@viewer_bp.route('/api/report/<int:report_id>', methods=['GET'])
@jwt_required()
def get_single_report(report_id):
    form_type = request.args.get('form_type', 'reporte_incidente')
    app_logger.info(f"Attempting to fetch single report with ID: {report_id} via GET /api/report/<id> with form_type: {form_type}")
    # fetch_reports_by_ids expects a list of IDs
    reports = fetch_reports_by_ids([report_id], form_type=form_type)
    
    if reports:
        app_logger.info(f"Successfully fetched report {report_id} for details.")
        return jsonify(reports[0]), 200
    else:
        app_logger.warning(f"Report with ID {report_id} not found for details.")
        return jsonify({"success": False, "message": f"Report with ID {report_id} not found."}), 404


@viewer_bp.route('/api/email-reports', methods=['POST'])
@jwt_required()
def email_selected_reports_api():
    user_email = get_jwt_identity()
    data = request.get_json()
    requests_payload  = data.get('reports')         # New format: list of {id, formType}
    report_ids        = data.get('report_ids')       # Old format: list of ids (ints)
    recipient_email   = data.get('recipient_email')  # Recipient address

    if not requests_payload and not report_ids:
        return jsonify({"success": False, "message": "No reports provided."}), 400

    if not recipient_email or not re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", recipient_email):
        return jsonify({"success": False, "message": "Invalid recipient email address."}), 400

    app_logger.info(f"User {user_email} requested to email reports to {recipient_email}")

    reports_to_email = []
    if requests_payload:
        reports_to_email = fetch_reports_mixed(requests_payload)
    elif report_ids:
        # Fallback for backward compatibility
        reports_to_email = fetch_reports_by_ids(report_ids)

    if not reports_to_email:
        app_logger.warning(f"No reports found for the provided items during email request.")
        return jsonify({"success": False, "message": "No reports found for the provided IDs."}), 404

    from admin_bp import get_operation_timezone, format_local_datetime
    tz = get_operation_timezone(reports=reports_to_email)

    form_types = {r.get('formType') for r in reports_to_email if r.get('formType')}
    if len(form_types) == 1:
        f_type = list(form_types)[0]
        cfg = FORM_CONFIGS.get(f_type, {})
        email_subtitle = cfg.get('title_prefix') or f"Reporte de {f_type.replace('_', ' ').title()}"
        subject = f"{email_subtitle} — Kanan Sentinel SekApp"
    elif len(form_types) > 1:
        email_subtitle = "Reporte Consolidado de Operaciones"
        subject = f"Reporte Consolidado — Kanan Sentinel SekApp"
    else:
        email_subtitle = "Reporte Operativo"
        subject = f"Reporte Operativo — Kanan Sentinel SekApp"

    SKIP_KEYS = {'URLs de Imágenes o PDFs', 'foto_evidencia_url', 'Foto Evidencia', 'Anexos', 'Latitude', 'Longitude'}
    logo_src = _get_logo_data_url() or ""
    logo_tag = f'<img src="{logo_src}" alt="Kanan" height="40" style="width:auto;border-radius:6px;background:transparent;padding:3px;vertical-align:middle;">' if logo_src else ''
    gen_date = format_local_datetime(datetime.now(), tz=tz, time_sep=" a las ")

    p = []
    p.append(f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f3f4f6;padding:24px 0;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:8px;overflow:hidden;border:1px solid #e2e8f0;">

  <!-- Header -->
  <tr>
    <td style="background:#1e3a8a;padding:16px 24px;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td style="vertical-align:middle;">{logo_tag}&nbsp;&nbsp;<span style="color:#ffffff;font-size:15px;font-weight:bold;vertical-align:middle;">Kanan Sentinel SekApp</span><br>
              <span style="color:#bfdbfe;font-size:11px;margin-left:48px;">{email_subtitle}</span></td>
          <td align="right" style="color:#bfdbfe;font-size:11px;vertical-align:middle;">Generado el<br>{gen_date}</td>
        </tr>
      </table>
    </td>
  </tr>

  <!-- Intro -->
  <tr>
    <td style="padding:20px 24px 8px 24px;color:#374151;font-size:13px;">
      <p style="margin:0 0 8px 0;">Estimado/a,</p>
      <p style="margin:0;color:#6b7280;">A continuación se presentan los detalles del reporte seleccionado.</p>
    </td>
  </tr>
""")

    for report in reports_to_email:
        data = report.get('data', {})
        foto_items = []  # list of (label, url)
        signatures = []  # list of (label, data_url)
        attachment_urls = []

        # Build optional map thumbnail for this report
        email_map_td = ''
        try:
            _lat = float(data.get('Latitude', '') or '')
            _lng = float(data.get('Longitude', '') or '')
            map_thumb = _map_thumbnail_html(_lat, _lng, width=150, height=90, clickable=True)
            email_map_td = (f'<td style="padding:0;width:160px;text-align:right;vertical-align:middle;">'
                            f'{map_thumb}</td>')
        except (ValueError, TypeError):
            pass

        report_date_str = format_local_datetime(report.get('dateSubmitted'), tz=tz, time_sep=" a las ")
        p.append(f"""
  <!-- Report card -->
  <tr><td style="padding:16px 24px 0 24px;">
    <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e2e8f0;border-radius:6px;overflow:hidden;">
      <!-- Report title bar -->
      <tr>
        <td colspan="2" style="background:#1d4ed8;padding:0;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td style="padding:10px 14px;vertical-align:middle;">
                <span style="color:#ffffff;font-size:13px;font-weight:bold;">{report['title']}</span><br>
                <span style="color:#bfdbfe;font-size:10px;">Enviado por: {report['submittedBy']} &nbsp;&middot;&nbsp; {report_date_str}</span>
              </td>
              {email_map_td}
            </tr>
          </table>
        </td>
      </tr>
""")

        row_idx = 0
        for key, value in data.items():
            if _is_blank_export_value(value):
                continue
            val_str_raw = str(value).strip() if value is not None else ""
            k_lower = key.lower()
            is_foto_key = any(t in k_lower for t in ('foto', 'evidencia', 'imagen', 'photo', 'anexo', 'urls de imágenes'))
            is_url_val = (val_str_raw.startswith('http://') or val_str_raw.startswith('https://') or
                          val_str_raw.startswith('/api/media'))

            if key in SKIP_KEYS or (is_foto_key and is_url_val):
                for url in val_str_raw.split('\n'):
                    url = url.strip()
                    if not url or _is_blank_export_value(url):
                        continue
                    if is_foto_key and key not in SKIP_KEYS:
                        foto_items.append((key, url))
                    else:
                        attachment_urls.append(url)
                continue
            if key.lower() == 'inventario':
                inv_table_html = _render_inventario_html_table(value, is_email=True)
                if inv_table_html:
                    bg = '#f8fafc' if row_idx % 2 == 0 else '#ffffff'
                    p.append(f"""      <tr style="background:{bg};">
        <td colspan="2" style="padding:10px 12px;border-bottom:1px solid #f1f5f9;background:#fafafa;">
          <strong style="font-size:11px;color:#374151;display:block;margin-bottom:6px;">Inventario:</strong>
          {inv_table_html}
        </td>
      </tr>
""")
                    row_idx += 1
                    continue

            if key == 'Lista Asistencia':
                lista_table_html = _render_lista_asistencia_html(value)
                if lista_table_html:
                    bg = '#f8fafc' if row_idx % 2 == 0 else '#ffffff'
                    p.append(f"""      <tr style="background:{bg};">
        <td colspan="2" style="padding:10px 12px;border-bottom:1px solid #f1f5f9;background:#fafafa;">
          <strong style="font-size:11px;color:#374151;display:block;margin-bottom:6px;">Lista de Asistencia:</strong>
          {lista_table_html}
        </td>
      </tr>
""")
                    row_idx += 1
                    continue

            val_str = str(value).strip()
            if 'firma' in key.lower() or val_str.startswith('data:image'):
                try:
                    sig_list = json.loads(val_str) if val_str.startswith('[') else None
                except Exception:
                    sig_list = None
                if isinstance(sig_list, list):
                    for i, sv in enumerate(sig_list):
                        sv = str(sv).strip()
                        if sv and sv not in ('N/A', 'None', ''):
                            label = f"{key} {i+1}" if len(sig_list) > 1 else key
                            signatures.append((label, sv))
                else:
                    signatures.append((key, val_str))
                continue
            bg = '#f8fafc' if row_idx % 2 == 0 else '#ffffff'
            clean_val = str(escape(val_str)).replace('\n', '<br>')
            p.append(f"""      <tr style="background:{bg};">
        <td style="padding:6px 12px;font-size:11px;font-weight:bold;color:#374151;width:38%;border-bottom:1px solid #f1f5f9;">{escape(key)}</td>
        <td style="padding:6px 12px;font-size:11px;color:#1f2937;border-bottom:1px solid #f1f5f9;">{clean_val}</td>
      </tr>
""")
            row_idx += 1

        # Photos row(s)
        if foto_items:
            fotos_html = ''.join(
                f'<div style="display:inline-block;margin-right:12px;margin-bottom:8px;vertical-align:top;text-align:center;border:1px solid #e2e8f0;border-radius:4px;padding:6px;background:#f8fafc;">'
                f'<p style="margin:0 0 4px 0;font-size:10px;font-weight:bold;color:#1e3a8a;">{escape(lbl)}</p>'
                f'<a href="{_media_proxy_url(u)}" target="_blank"><img src="{u}" alt="{escape(lbl)}" style="max-width:160px;max-height:120px;border-radius:3px;display:block;margin:0 auto;"></a>'
                f'</div>'
                for lbl, u in foto_items
            )
            p.append(f"""      <tr>
        <td colspan="2" style="padding:12px 14px;border-top:1px solid #e5e7eb;">
          <p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#1e3a8a;">Registro Fotográfico / Evidencias</p>
          {fotos_html}
        </td>
      </tr>
""")

        # Signature row(s)
        if signatures:
            sigs_html = ''.join(
                f'<div style="display:inline-block;margin-right:16px;vertical-align:top;">'
                f'<p style="margin:0 0 4px 0;font-size:11px;font-weight:bold;color:#374151;">{lbl}</p>'
                f'<img src="{sv}" alt="{lbl}" style="max-width:200px;max-height:90px;border:1px solid #d1d5db;border-radius:4px;padding:4px;background:#fff;">'
                f'</div>'
                for lbl, sv in signatures
            )
            p.append(f"""      <tr>
        <td colspan="2" style="padding:12px 14px;border-top:1px solid #e5e7eb;">
          {sigs_html}
        </td>
      </tr>
""")

        # Attachments
        if attachment_urls:
            p.append('<tr><td colspan="2" style="padding:12px 14px;border-top:1px solid #e5e7eb;"><p style="margin:0 0 8px 0;font-size:11px;font-weight:bold;color:#374151;">Archivos Adjuntos</p>')
            for url in attachment_urls:
                fname = url.split('?')[0].split('/')[-1]
                proxy_url = _media_proxy_url(url)
                lower = url.lower().split('?')[0]
                if lower.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                    p.append(f'<a href="{proxy_url}" target="_blank"><img src="{url}" alt="{fname}" style="max-width:180px;max-height:150px;border:1px solid #d1d5db;border-radius:4px;margin:4px 4px 0 0;"></a>')
                else:
                    p.append(f'<p style="margin:4px 0;font-size:11px;"><a href="{proxy_url}" target="_blank" style="color:#2563eb;">{fname}</a></p>')
            p.append('</td></tr>')

        p.append('    </table>\n  </td></tr>')

    # Footer
    p.append(f"""
  <tr>
    <td style="padding:20px 24px;border-top:1px solid #e5e7eb;color:#9ca3af;font-size:10px;text-align:center;">
      Generado por <strong>{user_email}</strong> desde Kanan Sentinel SekApp.<br>
      Este correo es generado automáticamente, por favor no responder.
    </td>
  </tr>

</table>
</td></tr></table>
</body></html>""")

    email_html_body = ''.join(p)

    success, message = send_reports_email(recipient_email, subject, email_html_body, is_html=True)

    if success:
        return jsonify({"success": True, "message": "Reportes enviados por correo electrónico exitosamente!"}), 200
    else:
        app_logger.error(f"Failed to send email: {message}")
        return jsonify({"success": False, "message": f"Error al enviar correo electrónico: {message}"}), 500


@viewer_bp.route('/api/export-excel', methods=['POST'])
@jwt_required()
def export_excel():
    """Export selected reports to Excel format"""
    user_email = get_jwt_identity()
    data = request.get_json()
    requests_payload = data.get('reports') # New format: list of {id, formType}
    report_ids = data.get('report_ids')

    if not requests_payload and not report_ids:
        return jsonify({"success": False, "message": "No reports provided."}), 400

    app_logger.info(f"User {user_email} requested Excel export")

    try:
        # Import openpyxl for Excel generation
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            app_logger.error("openpyxl not installed. Please install with: pip install openpyxl")
            return jsonify({"success": False, "message": "Excel export not available. Missing required dependencies."}), 500

        # Fetch the reports
        reports_to_export = []
        if requests_payload:
             reports_to_export = fetch_reports_mixed(requests_payload, skip_signing=True)
        elif report_ids:
             reports_to_export = fetch_reports_by_ids(report_ids, skip_signing=True)

        if not reports_to_export:
            app_logger.warning(f"No reports found for the provided IDs during Excel export.")
            return jsonify({"success": False, "message": "No reports found for the provided IDs."}), 404

        # Create Excel workbook
        wb = Workbook()
        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"), 
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Group reports by formType
        reports_by_type = {}
        for report in reports_to_export:
            f_type = report.get('formType', 'reporte_incidente')
            if f_type not in reports_by_type:
                reports_by_type[f_type] = []
            reports_by_type[f_type].append(report)

        # Create a sheet for each form type
        for f_type, type_reports in reports_by_type.items():
            # Get config for this form type to determine headers
            config = FORM_CONFIGS.get(f_type)
            if not config:
                app_logger.warning(f"No configuration found for form type: {f_type}. Skipping.")
                continue

            # Create sheet
            # sheet_title existe donde el nombre oficial no sobrevive al corte de
            # 30 caracteres de Excel sin volverse ambiguo frente a otra hoja.
            sheet_title = (config.get('sheet_title') or config.get('title_prefix', f_type))[:30]
            ws = wb.create_sheet(title=sheet_title)

            # Define headers dynamically based on data_mapping
            # Standard headers first
            headers = ["ID Reporte", "Enviado Por", "Fecha Envío"]
            
            # Dynamic headers from the mapping, minus the columns this form never
            # fills: a label left blank by every selected record is a leftover of
            # the table's legacy columns, not a field of the form being exported.
            dynamic_headers = [
                label for label in config['data_mapping']
                if any(not _is_blank_export_value(r.get('data', {}).get(label))
                       for r in type_reports)
            ]
            headers.extend(dynamic_headers)

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Write data
            from admin_bp import get_operation_timezone, format_local_datetime
            tz = get_operation_timezone(reports=reports_to_export)

            for row, report in enumerate(type_reports, 2):
                # Standard data
                ws.cell(row=row, column=1, value=report['id']).border = border
                ws.cell(row=row, column=2, value=report['submittedBy']).border = border
                date_sub_local = format_local_datetime(report.get('dateSubmitted'), tz=tz, time_sep=" ")
                ws.cell(row=row, column=3, value=date_sub_local).border = border

                # Dynamic data
                max_row_height = 25 # Default height
                
                for i, header_key in enumerate(dynamic_headers):
                    # Map header key to data key using config
                    data_key = header_key # The keys in report['data'] match the keys in data_mapping (labels)
                    val = report['data'].get(data_key, '')
                    if _is_blank_export_value(val):
                        val = ''
                    
                    col_index = 4 + i
                    if isinstance(val, (datetime, date)):
                        cell_value = format_local_datetime(val, tz=tz, time_sep=" ")
                    elif isinstance(val, (list, dict)):
                        cell_value = _format_structured_value_as_text(val)
                    elif isinstance(val, str) and any(k in header_key.lower() for k in ('fecha', 'fecha/hora', 'fecha hora', 'fecha evento', 'fecha incidente', 'fecha visita', 'fecha cumplimiento')):
                        cell_value = format_local_datetime(val, tz=tz, time_sep=" ")
                    else:
                        cell_value = str(val) if val is not None else ''
                    cell = ws.cell(row=row, column=col_index, value=cell_value)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                    # Image Embedding Logic
                    is_image_field = any(keyword in header_key.lower() for keyword in ['firma', 'foto', 'evidencia', 'diagrama', 'imagen'])
                    
                    if is_image_field and val and isinstance(val, str):
                        try:
                            from openpyxl.drawing.image import Image as OpenpyxlImage
                            from PIL import Image as PILImage
                            import base64
                            
                            img_file = None
                            
                            # Case A: Base64 Data URI
                            if val.strip().startswith('data:image'):
                                try:
                                    # Format: data:image/png;base64,.....
                                    header, encoded = val.strip().split(',', 1)
                                    img_bytes = base64.b64decode(encoded)
                                    
                                    # Convert to PNG if needed (for WebP or other formats)
                                    pil_img = PILImage.open(BytesIO(img_bytes))
                                    if pil_img.format not in ['PNG', 'JPEG', 'JPG', 'GIF', 'BMP']:
                                        png_buffer = BytesIO()
                                        pil_img.convert('RGB').save(png_buffer, format='PNG')
                                        png_buffer.seek(0)
                                        img_file = png_buffer
                                    else:
                                        img_file = BytesIO(img_bytes)
                                    
                                    cell.value = "Firma Digital" # Set friendly text
                                    cell.hyperlink = None # No external link for base64
                                except Exception as e:
                                    app_logger.error(f"Error processing base64 image: {e}")
                                    
                            # Case B: URL (HTTP/GCS)
                            elif 'http' in val or 'storage.googleapis.com' in val:
                                # Handle multiple images (newlines)
                                urls = val.split('\n')
                                
                                valid_url = None
                                for u in urls:
                                    if u.strip():
                                        valid_url = u.strip()
                                        break
                                
                                if valid_url:
                                    # Just set the hyperlink, do not attempt to download or embed
                                    cell.value = "Ver Imagen Original"
                                    cell.hyperlink = valid_url
                                    cell.style = "Hyperlink"

                            # If we successfully got an image file, resize and place it
                            if img_file:
                                img = OpenpyxlImage(img_file)
                                
                                # Resize image (thumbnail)
                                # Target height ~80px
                                aspect_ratio = img.width / img.height
                                new_height = 80
                                new_width = int(new_height * aspect_ratio)
                                
                                img.height = new_height
                                img.width = new_width
                                
                                # Anchor to cell
                                col_letter = get_column_letter(col_index)
                                cell_address = f"{col_letter}{row}"
                                img.anchor = cell_address
                                
                                ws.add_image(img)
                                
                                # Update max row height needed
                                max_row_height = max(max_row_height, 90)

                        except Exception as e:
                            app_logger.error(f"Error embedding image for field {header_key}: {e}")
                            # Keep original text value
                            pass

                # Set row height
                ws.row_dimensions[row].height = max_row_height

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                # Check header length
                max_length = max(max_length, len(headers[col-1]))
                # Check data lengths (sample first 50 rows)
                for row in range(2, min(52, ws.max_row + 1)):
                    cell_value = ws[f"{column_letter}{row}"].value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set header row height
            ws.row_dimensions[1].height = 25

        # Create filename
        custom_filename = data.get('filename')
        if custom_filename:
            # Sanitize filename
            custom_filename = re.sub(r'[^\w\-_]', '', custom_filename)
            filename = f"{custom_filename}.xlsx"
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"reportes_incidencias_{timestamp}.xlsx"

        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        app_logger.info(f"Excel file generated successfully for {len(reports_to_export)} reports")

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        app_logger.error(f"Error generating Excel file: {e}", exc_info=True)
        return jsonify({"success": False, "message": "Error generating Excel file"}), 500

@viewer_bp.route('/api/generate-pdf', methods=['POST'])
@jwt_required()
def generate_pdf():
    """Generate PDF for selected reports using WeasyPrint"""
    user_email = get_jwt_identity()
    data = request.get_json()
    requests_payload = data.get('reports') # New format: list of {id, formType}
    report_ids = data.get('report_ids')

    if not requests_payload and not report_ids:
        return jsonify({"success": False, "message": "No reports provided."}), 400

    # Sin las librerías nativas de WeasyPrint no hay PDF que entregar. Antes se
    # devolvía un PDF vacío hardcodeado: el usuario se llevaba un archivo que no
    # abre y nadie se enteraba de la falla. Mismo criterio que el endpoint del
    # Morning Briefing, que ya respondía 503 en este caso.
    if not WEASYPRINT_AVAILABLE:
        app_logger.error("PDF solicitado pero WeasyPrint no está disponible en este entorno")
        return jsonify({"success": False,
                        "message": "La generación de PDF no está disponible en este entorno."}), 503

    app_logger.info(f"User {user_email} requested PDF generation")

    try:
        # Fetch the reports
        reports_to_pdf = []
        if requests_payload:
            reports_to_pdf = fetch_reports_mixed(requests_payload)
        elif report_ids:
            reports_to_pdf = fetch_reports_by_ids(report_ids)

        if not reports_to_pdf:
            app_logger.warning(f"No reports found for the provided IDs during PDF request.")
            return jsonify({"success": False, "message": "No reports found for the provided IDs."}), 404

        # Generate HTML content for PDF
        html_content = generate_reports_html(reports_to_pdf)
        
        # Create PDF using WeasyPrint
        pdf_buffer = BytesIO()
        HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)

        # Create filename
        from admin_bp import get_operation_timezone
        tz = get_operation_timezone(reports=reports_to_pdf)
        custom_filename = data.get('filename')
        if custom_filename:
            # Sanitize filename
            custom_filename = re.sub(r'[^\w\-_]', '', custom_filename)
            filename = f"{custom_filename}.pdf"
        else:
            timestamp = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
            form_types = {r.get('formType') for r in reports_to_pdf if r.get('formType')}
            if len(form_types) == 1:
                prefix = list(form_types)[0]
            else:
                prefix = "reportes_operativos"
            filename = f"{prefix}_{timestamp}.pdf"

        app_logger.info(f"PDF generated successfully for {len(reports_to_pdf)} reports using WeasyPrint")

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )

    except Exception as e:
        app_logger.error(f"Error generating PDF: {e}", exc_info=True)
        return jsonify({"success": False, "message": "Error generating PDF"}), 500


_LOGO_DATA_URL_CACHE = None

def _get_logo_data_url():
    global _LOGO_DATA_URL_CACHE
    if _LOGO_DATA_URL_CACHE:
        return _LOGO_DATA_URL_CACHE
    try:
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo_full.png')
        with open(logo_path, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode()
            _LOGO_DATA_URL_CACHE = f"data:image/png;base64,{b64}"
            return _LOGO_DATA_URL_CACHE
    except Exception:
        pass
    return None


def _map_thumbnail_html(lat: float, lng: float, width: int = 160, height: int = 96, zoom: int = 17, clickable: bool = True) -> str:
    """Return an HTML snippet showing a 2×2 OSM tile thumbnail centered on lat/lng."""
    n = 2 ** zoom
    lat_rad = math.radians(lat)
    frac_x = (lng + 180) / 360 * n
    frac_y = (1 - math.log(math.tan(lat_rad) + 1 / math.cos(lat_rad)) / math.pi) / 2 * n
    tile_x = int(frac_x - 0.5)
    tile_y = int(frac_y - 0.5)
    pix_x = round((frac_x - tile_x) * 256)
    pix_y = round((frac_y - tile_y) * 256)
    left = width // 2 - pix_x
    top  = height // 2 - pix_y
    maps_url = f"https://www.google.com/maps?q={lat},{lng}&z=19"

    tiles = ''.join(
        f'<img src="https://tile.openstreetmap.org/{zoom}/{tile_x+dx}/{tile_y+dy}.png"'
        f' style="position:absolute;left:{dx*256}px;top:{dy*256}px;width:256px;height:256px;">'
        for dy in range(2) for dx in range(2)
    )
    pin = (f'<div style="position:absolute;left:{pix_x}px;top:{pix_y}px;'
           f'transform:translate(-50%,-100%);font-size:18px;line-height:1;'
           f'filter:drop-shadow(0 1px 3px rgba(0,0,0,0.7));">&#128205;</div>')
    inner = (f'<div style="position:absolute;width:512px;height:512px;left:{left}px;top:{top}px;">'
             f'{tiles}{pin}</div>')
    container_style = (f'position:relative;width:{width}px;height:{height}px;overflow:hidden;'
                       f'border-radius:4px;border:1px solid #d1d5db;display:inline-block;'
                       f'vertical-align:middle;flex-shrink:0;')
    if clickable:
        return (f'<a href="{maps_url}" target="_blank" rel="noopener noreferrer"'
                f' style="{container_style}text-decoration:none;">{inner}</a>')
    return f'<div style="{container_style}">{inner}</div>'


def _render_lista_asistencia_html(value):
    """Render a lista_asistencia JSON string as an HTML table for PDF."""
    try:
        attendees = json.loads(value) if isinstance(value, str) else value
    except Exception:
        return None
    if not isinstance(attendees, list) or not attendees:
        return None

    rows = []
    for a in attendees:
        if not isinstance(a, dict):
            continue
        nombre = a.get('nombre', '')
        cargo = a.get('cargo', '')
        num_emp = a.get('numero_empleado', '')
        doc = a.get('documento', '')
        firma = a.get('firma', '')
        via = a.get('via', '')

        firma_html = ''
        if firma and firma.startswith('data:image'):
            firma_html = f'<img src="{firma}" style="max-width:90px;max-height:45px;border:1px solid #d1d5db;border-radius:3px;object-fit:contain;">'
        elif via == 'QR':
            firma_html = '<span style="font-size:7pt;color:#6b7280;">QR</span>'

        rows.append(
            f'<tr>'
            f'<td style="padding:3px 6px;border-bottom:1px solid #e5e7eb;font-size:7.5pt;">{nombre}</td>'
            f'<td style="padding:3px 6px;border-bottom:1px solid #e5e7eb;font-size:7.5pt;">{cargo}</td>'
            f'<td style="padding:3px 6px;border-bottom:1px solid #e5e7eb;font-size:7.5pt;">{num_emp}</td>'
            f'<td style="padding:3px 6px;border-bottom:1px solid #e5e7eb;font-size:7.5pt;">{doc}</td>'
            f'<td style="padding:3px 6px;border-bottom:1px solid #e5e7eb;text-align:center;">{firma_html}</td>'
            f'</tr>'
        )

    if not rows:
        return None

    header = (
        '<tr style="background:#f1f5f9;">'
        '<th style="padding:4px 6px;font-size:7.5pt;text-align:left;border-bottom:1px solid #d1d5db;">Nombre</th>'
        '<th style="padding:4px 6px;font-size:7.5pt;text-align:left;border-bottom:1px solid #d1d5db;">Cargo</th>'
        '<th style="padding:4px 6px;font-size:7.5pt;text-align:left;border-bottom:1px solid #d1d5db;">N° Empleado</th>'
        '<th style="padding:4px 6px;font-size:7.5pt;text-align:left;border-bottom:1px solid #d1d5db;">Documento</th>'
        '<th style="padding:4px 6px;font-size:7.5pt;text-align:center;border-bottom:1px solid #d1d5db;">Firma</th>'
        '</tr>'
    )
    return (
        f'<table style="width:100%;border-collapse:collapse;border:1px solid #e2e8f0;border-radius:3px;">'
        f'<thead>{header}</thead>'
        f'<tbody>{"".join(rows)}</tbody>'
        f'</table>'
    )


_VISITA_ESTADO_LABELS = {
    'cumplido':   'Cumplido',
    'cerrado':    'Cumplido',
    'completado': 'Cumplido',
    'finalizado': 'Cumplido',
    'ejecutado':  'Cumplido',
    'vencido':    'Vencido',
    'pendiente':  'Pendiente',
}

_VISITA_ESTADO_COLORS = {
    'Cumplido':  '#15803d',
    'Pendiente': '#b45309',
    'Vencido':   '#b91c1c',
}

# Fields of the acta that are folded into the single "Acuerdos y Compromisos" table.
_VISITA_COMPROMISO_KEYS = (
    'Temas Tratados', 'Acuerdos', 'Nombre Responsable',
    'Compromisos Estados', 'Fecha Cumplimiento',
)


def _visita_clean_text(val):
    """Return a trimmed string, treating the viewer's placeholders as empty."""
    text = str(val).strip() if val is not None else ''
    return '' if text in ('N/A', 'None') else text


def _visita_parse_fecha(raw):
    """Parse an ISO-ish date string into a date, or None."""
    text = _visita_clean_text(raw)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        return None


def _visita_fecha_legible(raw):
    """Format a date as dd/mm/yyyy for the document; fall back to the raw text."""
    parsed = _visita_parse_fecha(raw)
    if parsed:
        return parsed.strftime('%d/%m/%Y')
    return _visita_clean_text(raw)


def _render_participantes_visita_html(value):
    """Render detalles_participantes as a Nombre/Cargo/Firma table for PDF."""
    participantes = _ensure_json_serializable(value)
    if isinstance(participantes, dict):
        participantes = [participantes]
    if not isinstance(participantes, list):
        return None

    rows = []
    for item in participantes:
        # Participant entries are structured records.  Never print an unknown
        # scalar/object representation because it may contain internal data or a
        # complete base64 signature string.
        if not isinstance(item, dict):
            continue
        nombre = _visita_clean_text(item.get('nombre'))
        cargo  = _visita_clean_text(item.get('cargo'))
        firma  = _visita_clean_text(item.get('firma'))

        if not (nombre or cargo or firma):
            continue

        firma_src = ''
        if firma.startswith('data:image'):
            firma_src = firma
        elif firma.startswith(('https://', 'http://')):
            firma_src = generate_signed_url(firma)

        if firma_src:
            firma_html = ('<img src="' + str(escape(firma_src)) + '" style="max-width:110px;max-height:50px;'
                          'border:1px solid #d1d5db;border-radius:3px;object-fit:contain;">')
        else:
            firma_html = '<span style="font-size:7pt;color:#6b7280;">Sin firma</span>'

        td = 'padding:3px 6px;border-bottom:1px solid #e5e7eb;font-size:7.5pt;'
        rows.append(
            '<tr>'
            '<td style="' + td + '">' + str(escape(nombre or '—')) + '</td>'
            '<td style="' + td + '">' + str(escape(cargo or '—')) + '</td>'
            '<td style="' + td + 'text-align:center;">' + firma_html + '</td>'
            '</tr>'
        )

    if not rows:
        return None

    th = 'padding:4px 6px;font-size:7.5pt;text-align:left;border-bottom:1px solid #d1d5db;'
    header = (
        '<tr style="background:#f1f5f9;">'
        '<th style="' + th + '">Nombre</th>'
        '<th style="' + th + '">Cargo</th>'
        '<th style="' + th + 'text-align:center;">Firma</th>'
        '</tr>'
    )
    return ('<table style="width:100%;border-collapse:collapse;border:1px solid #e2e8f0;border-radius:3px;">'
            '<thead>' + header + '</thead><tbody>' + ''.join(rows) + '</tbody></table>')


def _render_compromisos_visita_html(data):
    """Render the acta's temas/acuerdos/responsables/estados as one readable table.

    Reuses the same parsing and status precedence as the Gestión dashboard so the
    PDF and the dashboard never disagree about a commitment's state.
    """
    from dashboard_bp import _visita_split_blocks, _visita_parse_responsables, _visita_status

    temas    = _visita_split_blocks(_visita_clean_text(data.get('Temas Tratados')))
    acuerdos = _visita_split_blocks(_visita_clean_text(data.get('Acuerdos')))

    raw_resp = data.get('Nombre Responsable')
    if isinstance(raw_resp, str) and not _visita_clean_text(raw_resp):
        raw_resp = None
    responsables = _visita_parse_responsables(raw_resp)

    estados_override = _ensure_json_serializable(data.get('Compromisos Estados'))
    if not isinstance(estados_override, dict):
        estados_override = {}

    fecha_general = _visita_clean_text(data.get('Fecha Cumplimiento'))
    hoy = datetime.now().date()

    total = max(len(temas), len(acuerdos), len(responsables))
    if not total:
        return None

    show_tema = any(temas)
    filas = []
    for idx in range(total):
        tema    = temas[idx] if idx < len(temas) else ''
        acuerdo = acuerdos[idx] if idx < len(acuerdos) else ''
        item    = responsables[idx] if idx < len(responsables) and isinstance(responsables[idx], dict) else {}

        nombre = _visita_clean_text(item.get('nombre'))
        if nombre.startswith('[') or nombre.startswith('{'):
            nombre = ''

        fecha_raw  = _visita_clean_text(item.get('fecha')) or fecha_general
        fecha_dt   = _visita_parse_fecha(fecha_raw)
        fecha_text = _visita_fecha_legible(fecha_raw)

        override = estados_override.get(str(idx))
        if isinstance(override, dict):
            raw_estado = _visita_clean_text(override.get('estado'))
        else:
            raw_estado = _visita_clean_text(override)
        if not raw_estado:
            raw_estado = _visita_clean_text(item.get('estado'))
        if not raw_estado:
            raw_estado = _visita_status(acuerdo, fecha_dt)

        estado = _VISITA_ESTADO_LABELS.get(raw_estado.lower(), raw_estado.capitalize())
        if estado == 'Pendiente' and fecha_dt and fecha_dt < hoy:
            estado = 'Vencido'

        color = _VISITA_ESTADO_COLORS.get(estado, '#374151')
        td = 'padding:4px 6px;border-bottom:1px solid #e5e7eb;font-size:7.5pt;vertical-align:top;'

        celdas = ['<td style="' + td + 'text-align:center;color:#6b7280;">' + str(idx + 1) + '</td>']
        if show_tema:
            celdas.append('<td style="' + td + '">' + str(escape(tema or '—')).replace('\n', '<br>') + '</td>')
        celdas.append('<td style="' + td + '">' + str(escape(acuerdo or '—')).replace('\n', '<br>') + '</td>')
        celdas.append('<td style="' + td + '">' + str(escape(nombre or '—')) + '</td>')
        celdas.append('<td style="' + td + 'white-space:nowrap;">' + str(escape(fecha_text or '—')) + '</td>')
        celdas.append('<td style="' + td + 'white-space:nowrap;font-weight:bold;color:' + color + ';">'
                      + str(escape(estado or '—')) + '</td>')
        filas.append('<tr>' + ''.join(celdas) + '</tr>')

    th = 'padding:4px 6px;font-size:7.5pt;text-align:left;border-bottom:1px solid #d1d5db;'
    encabezados = ['<th style="' + th + 'text-align:center;width:18px;">#</th>']
    if show_tema:
        encabezados.append('<th style="' + th + '">Tema tratado</th>')
    encabezados.append('<th style="' + th + '">Acuerdo / Compromiso</th>')
    encabezados.append('<th style="' + th + '">Responsable</th>')
    encabezados.append('<th style="' + th + '">Fecha compromiso</th>')
    encabezados.append('<th style="' + th + '">Estado</th>')

    return ('<table style="width:100%;border-collapse:collapse;border:1px solid #e2e8f0;border-radius:3px;">'
            '<thead><tr style="background:#f1f5f9;">' + ''.join(encabezados) + '</tr></thead>'
            '<tbody>' + ''.join(filas) + '</tbody></table>')


def _render_personas_involucradas_html(value):
    """Render personas_involucradas as a Tipo/Nombre table for PDF."""
    personas = _ensure_json_serializable(value)
    if isinstance(personas, dict):
        personas = [personas]
    if not isinstance(personas, list) or not personas:
        return None

    rows = []
    for item in personas:
        if isinstance(item, dict):
            tipo = _visita_clean_text(item.get('tipo') or item.get('persona_tipo'))
            nombre = _visita_clean_text(item.get('nombre') or item.get('persona_nombre'))
            if not (tipo or nombre):
                continue
            td = 'padding:3px 6px;border-bottom:1px solid #e5e7eb;font-size:7.5pt;'
            rows.append(
                '<tr>'
                '<td style="' + td + 'font-weight:600;width:35%;">' + str(escape(tipo or '—')) + '</td>'
                '<td style="' + td + '">' + str(escape(nombre or '—')) + '</td>'
                '</tr>'
            )
        elif isinstance(item, str) and item.strip():
            td = 'padding:3px 6px;border-bottom:1px solid #e5e7eb;font-size:7.5pt;'
            rows.append('<tr><td colspan="2" style="' + td + '">' + str(escape(item.strip())) + '</td></tr>')

    if not rows:
        return None

    th = 'padding:4px 6px;font-size:7.5pt;text-align:left;border-bottom:1px solid #d1d5db;'
    header = (
        '<tr style="background:#f1f5f9;">'
        '<th style="' + th + 'width:35%;">Tipo</th>'
        '<th style="' + th + '">Nombre completo</th>'
        '</tr>'
    )
    return ('<table style="width:100%;border-collapse:collapse;border:1px solid #e2e8f0;border-radius:3px;">'
            '<thead>' + header + '</thead><tbody>' + ''.join(rows) + '</tbody></table>')


def generate_reports_html(reports):
    """Generate HTML content for PDF generation."""
    from admin_bp import get_operation_timezone, format_local_datetime
    tz = get_operation_timezone(reports=reports)

    SKIP_KEYS = {'URLs de Imágenes o PDFs', 'foto_evidencia_url', 'Foto Evidencia', 'Anexos'}
    # Internal identifiers and raw coordinates: the PDF is handed to the client, so they
    # never appear. Kept apart from SKIP_KEYS, whose values are parsed as attachment URLs.
    HIDDEN_KEYS = {
        'Company Id', 'Customer Company Id', 'Id Propiedad',
        'Location Accuracy', 'Latitude', 'Longitude',
        'Cliente Instalacion', 'Submitter Timezone'
    }

    def _is_signature(key, val_str):
        return 'firma' in key.lower() or val_str.startswith('data:image')
    logo_src = _get_logo_data_url() or ""

    logo_img = f'<img src="{logo_src}" alt="">' if logo_src else '<span style="font-size:20pt;color:#1d4ed8;">&#9632;</span>'

    form_types = {r.get('formType') for r in reports if r.get('formType')}
    if len(form_types) == 1:
        f_type = list(form_types)[0]
        cfg = FORM_CONFIGS.get(f_type, {})
        header_subtitle = cfg.get('title_prefix') or cfg.get('sheet_title') or f"Reporte de {f_type.replace('_', ' ').title()}"
    elif len(form_types) > 1:
        header_subtitle = "Reporte Consolidado de Operaciones"
    else:
        header_subtitle = "Reporte Operativo SEKapp"

    gen_date_str = format_local_datetime(datetime.now(), tz=tz, time_sep=" a las ")

    # `escape()` devuelve un Markup, no un str, y Markup redefine la suma: al
    # concatenar `str + Markup` markupsafe escapa el operando izquierdo. En una
    # cadena que se evalúa de izquierda a derecha eso convierte TODO el documento
    # anterior —DOCTYPE, <head> y el bloque <style> completo— en texto escapado,
    # y WeasyPrint termina imprimiendo el CSS en vez de aplicarlo. De ahí el
    # str() alrededor de cada escape() que se concatene aquí abajo.
    html_parts = ["""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: Arial, Helvetica, sans-serif; color: #1f2937; font-size: 8.5pt; line-height: 1.4; }
.page-header {
    background: #1e3a8a;
    color: white;
    padding: 10px 16px;
    display: table;
    width: 100%;
}
.header-left { display: table-cell; vertical-align: middle; }
.header-logo { display: inline-block; vertical-align: middle; margin-right: 10px; }
.header-logo img { height: 40px; width: auto; border-radius: 6px; background: transparent; padding: 2px; vertical-align: middle; object-fit: contain; }
.header-title { display: inline-block; vertical-align: middle; }
.header-title h1 { font-size: 12pt; font-weight: bold; }
.header-title p { font-size: 8pt; opacity: 0.8; margin-top: 1px; }
.header-right { display: table-cell; vertical-align: middle; text-align: right; font-size: 7.5pt; opacity: 0.85; white-space: nowrap; }
.report-block {
    background: white;
    margin: 10px 12px;
    border: 1px solid #e2e8f0;
    border-radius: 6px;
    overflow: hidden;
}
.report-title-bar {
    background: #1d4ed8;
    color: white;
    padding: 8px 14px;
    display: table;
    width: 100%;
}
.report-title-bar h2 { font-size: 11pt; font-weight: bold; }
.report-title-bar .meta { font-size: 7.5pt; opacity: 0.85; margin-top: 2px; }
.report-body { padding: 10px 14px; }
table.fields { width: 100%; border-collapse: collapse; }
table.fields td { padding: 4px 8px; font-size: 8pt; border-bottom: 1px solid #f1f5f9; vertical-align: top; }
table.fields tr:last-child td { border-bottom: none; }
table.fields tr:nth-child(even) td { background: #f8fafc; }
td.lbl { width: 36%; font-weight: bold; color: #374151; white-space: nowrap; }
td.val { color: #1f2937; }
.bottom-row { display: table; width: 100%; margin-top: 8px; border-top: 1px solid #e5e7eb; padding-top: 8px; }
.sig-cell { display: table-cell; vertical-align: top; width: 50%; }
.att-cell { display: table-cell; vertical-align: top; width: 50%; padding-left: 12px; }
.section-label { font-size: 7.5pt; font-weight: bold; color: #374151; margin-bottom: 4px; }
.sig-img { max-width: 180px; max-height: 80px; border: 1px solid #d1d5db; border-radius: 3px; padding: 3px; object-fit: contain; }
.att-grid img { max-width: 120px; max-height: 90px; border: 1px solid #d1d5db; border-radius: 3px; object-fit: contain; }
.pdf-link { color: #2563eb; text-decoration: none; font-size: 8pt; }
.foto-section { margin-top: 10px; padding-top: 8px; border-top: 1px solid #e5e7eb; }
.foto-grid { width: 100%; }
.foto-item-cell { display: inline-block; margin-right: 10px; margin-bottom: 8px; vertical-align: top; text-align: center; border: 1px solid #d1d5db; border-radius: 4px; padding: 6px; background: #f8fafc; }
.foto-label { font-size: 7.5pt; font-weight: bold; color: #1e3a8a; margin-bottom: 4px; max-width: 170px; word-break: break-word; }
.foto-img { max-width: 170px; max-height: 125px; border-radius: 3px; object-fit: contain; display: block; margin: 0 auto; }
.map-thumb-cell { display: table-cell; vertical-align: middle; text-align: right; width: 170px; padding-left: 8px; }
.map-thumb-cell a { display: inline-block; border-radius: 4px; overflow: hidden; border: 1px solid rgba(255,255,255,0.35); }
.map-thumb-cell img { width: 150px; height: 90px; display: block; }
@page {
    size: A4;
    margin: 1cm 0.8cm;
    @bottom-center {
        content: "Página " counter(page) " — Kanan Sentinel SekApp";
        font-size: 7pt; color: #9ca3af;
    }
}
</style>
</head>
<body>
<div class="page-header">
    <div class="header-left">
        <span class="header-logo">""" + logo_img + """</span>
        <span class="header-title">
            <h1>Kanan Sentinel SekApp</h1>
            <p>""" + str(escape(header_subtitle)) + """</p>
        </span>
    </div>
    <div class="header-right">Generado el<br>""" + gen_date_str + """</div>
</div>
"""]

    for idx, report in enumerate(reports):
        data = report.get('data', {})
        pb = '' if idx == 0 else 'page-break-before:always;'

        # Build optional map thumbnail for title bar
        map_cell = ''
        try:
            _lat = float(data.get('Latitude', '') or '')
            _lng = float(data.get('Longitude', '') or '')
            maps_url = f"https://www.google.com/maps?q={_lat},{_lng}&z=19"
            map_thumb = _map_thumbnail_html(_lat, _lng, width=150, height=90, clickable=True)
            map_cell = f'<div class="map-thumb-cell">{map_thumb}</div>'
        except (ValueError, TypeError):
            pass

        report_date_str = format_local_datetime(report.get("dateSubmitted"), tz=tz, time_sep=" a las ")

        html_parts.append(
            f'<div class="report-block" style="{pb}">'
            f'<div class="report-title-bar" style="display:table;width:100%;">'
            f'<div style="display:table-cell;vertical-align:middle;">'
            f'<h2>{report["title"]}</h2>'
            f'<p class="meta">Enviado por: {report["submittedBy"]} &nbsp;&middot;&nbsp; {report_date_str}</p>'
            f'</div>'
            f'{map_cell}'
            f'</div>'
            f'<div class="report-body">'
            f'<table class="fields">'
        )

        foto_items = []  # list of (label, url)
        signatures = []  # list of (label, data_url)
        image_urls, pdf_urls, other_urls = [], [], []
        is_acta = report.get('formType') == 'registro_y_acta_de_visita'
        compromisos_rendered = False

        for key, value in data.items():
            if key in HIDDEN_KEYS:
                continue

            # Only what this record actually captured: columns the form no longer
            # asks for arrive as the "N/A" placeholder and used to print as rows
            # of a form that never had them.
            if _is_blank_export_value(value):
                continue

            val_str_raw = str(value).strip() if value is not None else ""
            k_lower = key.lower()

            is_foto_key = any(t in k_lower for t in ('foto', 'evidencia', 'imagen', 'photo', 'anexo', 'urls de imágenes'))
            is_url_val = (val_str_raw.startswith('http://') or val_str_raw.startswith('https://') or
                          val_str_raw.startswith('/api/media'))

            if key in SKIP_KEYS or (is_foto_key and is_url_val):
                # Parse attachment URLs
                for url in val_str_raw.split('\n'):
                    url = url.strip()
                    if not url or _is_blank_export_value(url):
                        continue
                    lower = url.lower().split('?')[0]
                    if lower.endswith(('.jpeg', '.jpg', '.png', '.gif', '.webp', '.svg')) or 'storage.googleapis.com' in lower or url.startswith('data:image'):
                        if is_foto_key and key not in SKIP_KEYS:
                            foto_items.append((key, url))
                        else:
                            image_urls.append(url)
                    elif lower.endswith('.pdf'):
                        pdf_urls.append(url)
                    else:
                        other_urls.append(url)
                continue

            if key.lower() == 'inventario':
                inv_table_html = _render_inventario_html_table(value, is_email=False)
                if inv_table_html:
                    html_parts.append(
                        f'<tr><td colspan="2" style="padding: 10px 8px; border-bottom: 1px solid #e2e8f0; background: #fafafa;">'
                        f'<strong style="color: #374151; font-size: 8pt; display: block; margin-bottom: 5px;">Inventario:</strong>'
                        f'{inv_table_html}'
                        f'</td></tr>'
                    )
                continue

            if key == 'Personas Involucradas':
                personas_html = _render_personas_involucradas_html(value)
                if personas_html:
                    html_parts.append(
                        f'<tr><td colspan="2" style="padding: 10px 8px; border-bottom: 1px solid #e2e8f0; background: #fafafa;">'
                        f'<strong style="color: #374151; font-size: 8pt; display: block; margin-bottom: 5px;">Personas Involucradas:</strong>'
                        f'{personas_html}'
                        f'</td></tr>'
                    )
                continue

            if key == 'Lista Asistencia':
                lista_html = _render_lista_asistencia_html(value)
                if lista_html:
                    html_parts.append(
                        f'<tr><td colspan="2" style="padding: 10px 8px; border-bottom: 1px solid #e2e8f0; background: #fafafa;">'
                        f'<strong style="color: #374151; font-size: 8pt; display: block; margin-bottom: 5px;">Lista Asistencia:</strong>'
                        f'{lista_html}'
                        f'</td></tr>'
                    )
                continue

            if key == 'Detalles Participantes':
                participantes_html = _render_participantes_visita_html(value)
                if participantes_html:
                    html_parts.append(
                        f'<tr><td colspan="2" style="padding: 10px 8px; border-bottom: 1px solid #e2e8f0; background: #fafafa;">'
                        f'<strong style="color: #374151; font-size: 8pt; display: block; margin-bottom: 5px;">Participantes:</strong>'
                        f'{participantes_html}'
                        f'</td></tr>'
                    )
                continue

            if is_acta and key in _VISITA_COMPROMISO_KEYS:
                # Rendered once, as a single table, in place of the raw per-field values.
                if not compromisos_rendered:
                    compromisos_rendered = True
                    compromisos_html = _render_compromisos_visita_html(data)
                    if compromisos_html:
                        html_parts.append(
                            f'<tr><td colspan="2" style="padding: 10px 8px; border-bottom: 1px solid #e2e8f0; background: #fafafa;">'
                            f'<strong style="color: #374151; font-size: 8pt; display: block; margin-bottom: 5px;">Acuerdos y Compromisos:</strong>'
                            f'{compromisos_html}'
                            f'</td></tr>'
                        )
                continue

            if isinstance(value, (datetime, date)):
                val_str = format_local_datetime(value, tz=tz, time_sep=" ")
            elif isinstance(value, str) and _es_clave_de_fecha(key):
                val_str = format_local_datetime(value, tz=tz, time_sep=" ")
            else:
                val_str = str(value).strip()

            if _is_signature(key, val_str):
                # val may be a JSON array of data URLs (multiple guards)
                try:
                    sig_list = json.loads(val_str) if val_str.startswith('[') else None
                except Exception:
                    sig_list = None
                if isinstance(sig_list, list):
                    for i, sv in enumerate(sig_list):
                        sv = str(sv).strip()
                        if _is_blank_export_value(sv):
                            continue
                        label = f"{key} {i+1}" if len(sig_list) > 1 else key
                        signatures.append((label, sv))
                else:
                    signatures.append((key, val_str))
                continue

            clean_value = str(escape(val_str)).replace('\n', '<br>')
            html_parts.append(f'<tr><td class="lbl">{escape(key)}</td><td class="val">{clean_value}</td></tr>')

        html_parts.append('</table>')

        # Photos / Evidence section
        if foto_items:
            html_parts.append(
                '<div class="foto-section">'
                '<p class="section-label" style="font-size:8pt;font-weight:bold;color:#1e3a8a;margin-bottom:6px;">Registro Fotográfico / Evidencias</p>'
                '<div class="foto-grid">'
            )
            for lbl, url in foto_items:
                clean_lbl = escape(lbl)
                html_parts.append(
                    f'<div class="foto-item-cell">'
                    f'<p class="foto-label">{clean_lbl}</p>'
                    f'<a href="{_media_proxy_url(url)}"><img class="foto-img" src="{url}" alt="{clean_lbl}"></a>'
                    f'</div>'
                )
            html_parts.append('</div></div>')

        # Signature + attachments side by side
        has_sig = bool(signatures)
        has_att = bool(image_urls or pdf_urls or other_urls)

        if has_sig or has_att:
            html_parts.append('<div class="bottom-row">')

            if has_sig:
                sigs_html = ''.join(
                    f'<div style="display:inline-block;margin-right:16px;vertical-align:top;">'
                    f'<p class="section-label">{lbl}</p>'
                    f'<img class="sig-img" src="{sv}" alt="{lbl}">'
                    f'</div>'
                    for lbl, sv in signatures
                )
                html_parts.append(f'<div class="sig-cell">{sigs_html}</div>')

            if has_att:
                html_parts.append('<div class="att-cell"><p class="section-label">Archivos Adjuntos</p><div class="att-grid">')
                for url in image_urls:
                    fname = url.split('?')[0].split('/')[-1]
                    html_parts.append(f'<a href="{_media_proxy_url(url)}"><img src="{url}" alt="{fname}"></a>')
                for url in pdf_urls:
                    fname = url.split('?')[0].split('/')[-1]
                    html_parts.append(f'<a href="{_media_proxy_url(url)}" class="pdf-link">&#128196; {fname}</a><br>')
                for url in other_urls:
                    fname = url.split('?')[0].split('/')[-1]
                    html_parts.append(f'<a href="{_media_proxy_url(url)}" class="pdf-link">&#128206; {fname}</a><br>')
                html_parts.append('</div></div>')

            html_parts.append('</div>')

        html_parts.append('</div></div>')

    html_parts.append('</body></html>')
    return ''.join(html_parts)


# ── Backup de Información ────────────────────────────────────────────────────

def _ensure_backups_table(conn):
    """Crea la tabla de backups si falta. Mismo patrón que asignaciones_hallazgo:
    la migración viaja con el código para no depender de correr schema.sql."""
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS backups_realizados (
                id               SERIAL PRIMARY KEY,
                generado_en      TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                generado_por     TEXT,
                periodo_desde    DATE,
                periodo_hasta    DATE,
                cliente_id       INTEGER,
                cliente_nombre   TEXT,
                propiedad_id     INTEGER,
                propiedad_nombre TEXT,
                total_registros  INTEGER NOT NULL DEFAULT 0,
                formato          TEXT,
                archivo          TEXT,
                company_id       INTEGER
            )
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_backups_generado_en
                ON backups_realizados(generado_en DESC)
        """)
    conn.commit()


def _backup_nombre_alcance(cur, cliente_id, propiedad_id):
    """Nombres legibles del alcance, para el archivo y para el registro."""
    cliente_nombre = propiedad_nombre = None
    if cliente_id:
        cur.execute("SELECT name FROM customer_companies WHERE id = %s", (cliente_id,))
        row = cur.fetchone()
        if row:
            cliente_nombre = row[0]
    if propiedad_id:
        cur.execute("SELECT nombre FROM propiedades WHERE id_propiedad = %s", (propiedad_id,))
        row = cur.fetchone()
        if row:
            propiedad_nombre = row[0]
    return cliente_nombre, propiedad_nombre


@viewer_bp.route('/api/backup', methods=['POST'])
@jwt_required()
def generar_backup():
    """
    Respalda TODO lo registrado en el período y alcance pedidos — los 11
    formularios, no lo que esté filtrado en pantalla.

    Devuelve un ZIP con dos vistas de lo mismo: un Excel multi-hoja para
    consultar y un JSON con los campos crudos, que es el que sirve para
    recuperar. Las URLs de GCS van sin firmar: una firma caduca y dejaría el
    backup con enlaces muertos.
    """
    import json as _json
    import zipfile
    from io import BytesIO as _BytesIO

    payload      = request.get_json() or {}
    desde        = (payload.get('start_date') or '').strip()
    hasta        = (payload.get('end_date') or '').strip()
    cliente_id   = payload.get('cliente') or None
    propiedad_id = payload.get('property_id') or None

    if not desde or not hasta:
        return jsonify({"success": False, "message": "Indique la fecha desde y la fecha hasta."}), 400
    try:
        d_desde = datetime.strptime(desde, '%Y-%m-%d').date()
        d_hasta = datetime.strptime(hasta, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"success": False, "message": "Formato de fecha inválido."}), 400
    if d_desde > d_hasta:
        return jsonify({"success": False, "message": "La fecha desde no puede ser posterior a la fecha hasta."}), 400

    user_email = get_jwt_identity()
    app_logger.info("Backup solicitado por %s: %s a %s (cliente=%s, propiedad=%s)",
                    user_email, desde, hasta, cliente_id, propiedad_id)

    filters = {'start_date': desde, 'end_date': hasta}
    if cliente_id:
        filters['cliente'] = str(cliente_id)
    if propiedad_id:
        filters['property_id'] = str(propiedad_id)

    conn = None
    try:
        # Sin límite práctico: un backup parcial no es un backup.
        registros, total = fetch_reports(0, 1000000, filters=filters,
                                         form_type='all', skip_signing=True)

        por_tipo = {}
        for r in registros:
            por_tipo.setdefault(r.get('formType', 'desconocido'), []).append(r)

        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "message": "Base de datos no disponible."}), 500
        _ensure_backups_table(conn)
        cur = conn.cursor()
        cliente_nombre, propiedad_nombre = _backup_nombre_alcance(cur, cliente_id, propiedad_id)

        alcance = cliente_nombre or 'Todos los clientes'
        if propiedad_nombre:
            alcance += f' - {propiedad_nombre}'
        sello = datetime.now().strftime('%Y%m%d_%H%M%S')
        base  = re.sub(r'[^\w\-]+', '_', f"backup_{alcance}_{desde}_{hasta}").strip('_')
        nombre_zip = f"{base}_{sello}.zip"

        manifiesto = {
            'generado_en':   datetime.now().isoformat(),
            'generado_por':  user_email,
            'periodo':       {'desde': desde, 'hasta': hasta},
            'alcance': {
                'cliente_id': cliente_id, 'cliente': cliente_nombre or 'Todos',
                'propiedad_id': propiedad_id, 'propiedad': propiedad_nombre or 'Todas',
            },
            'total_registros': len(registros),
            'registros_por_formulario': {k: len(v) for k, v in sorted(por_tipo.items())},
            'nota': ('Las URLs de evidencia apuntan a Google Cloud Storage sin firmar; '
                     'requieren credenciales del proyecto para descargarse.'),
        }

        buffer = _BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as z:
            z.writestr('manifiesto.json',
                       _json.dumps(manifiesto, ensure_ascii=False, indent=2, default=str))
            # Un JSON por formulario: recuperable y fácil de inspeccionar por partes.
            for f_type, filas in sorted(por_tipo.items()):
                z.writestr(f'datos/{f_type}.json',
                           _json.dumps(filas, ensure_ascii=False, indent=2, default=str))
            excel = _backup_excel_bytes(por_tipo)
            if excel:
                z.writestr(f'{base}.xlsx', excel)
        buffer.seek(0)

        cur.execute(
            """
            INSERT INTO backups_realizados
                (generado_por, periodo_desde, periodo_hasta, cliente_id, cliente_nombre,
                 propiedad_id, propiedad_nombre, total_registros, formato, archivo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'zip', %s)
            """,
            (user_email, d_desde, d_hasta, cliente_id, cliente_nombre,
             propiedad_id, propiedad_nombre, len(registros), nombre_zip)
        )
        conn.commit()
        cur.close()

        app_logger.info("Backup generado: %s (%d registros)", nombre_zip, len(registros))
        return send_file(buffer, as_attachment=True, download_name=nombre_zip,
                         mimetype='application/zip')

    except Exception as e:
        if conn:
            conn.rollback()
        app_logger.error(f"generar_backup error: {e}", exc_info=True)
        return jsonify({"success": False, "message": "Error interno al generar el backup."}), 500
    finally:
        if conn:
            conn.close()


def _backup_excel_bytes(por_tipo):
    """Hoja por formulario con las etiquetas de data_mapping. Vista legible del
    mismo contenido que el JSON; si openpyxl falta, el ZIP igual lleva el JSON."""
    try:
        from openpyxl import Workbook
    except ImportError:
        app_logger.warning("openpyxl no disponible: el backup sale solo con JSON")
        return None

    wb = Workbook()
    wb.remove(wb.active)
    for f_type, filas in sorted(por_tipo.items()):
        config = FORM_CONFIGS.get(f_type)
        titulo = (config.get('sheet_title') or config.get('title_prefix', f_type))[:30] if config else f_type[:30]
        ws = wb.create_sheet(title=titulo)

        etiquetas = []
        for fila in filas:
            for k in (fila.get('data') or {}):
                if k not in etiquetas:
                    etiquetas.append(k)
        ws.append(['ID', 'Formulario', 'Enviado Por', 'Fecha Envío'] + etiquetas)
        for fila in filas:
            datos = fila.get('data') or {}
            # Hora de operación, igual que el resto de las vistas y que el
            # export normal; `dateSubmitted` sigue en ISO solo para ordenar.
            ws.append([fila.get('id'), fila.get('title'), fila.get('submittedBy'),
                       fila.get('dateSubmittedLocal') or fila.get('dateSubmitted')] +
                      [_excel_safe(datos.get(e)) for e in etiquetas])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _excel_safe(v):
    """openpyxl solo acepta escalares; listas y dicts van serializados."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return '' if v is None else v
    return str(v)


@viewer_bp.route('/api/backup/estado', methods=['GET'])
@jwt_required()
def backup_estado():
    """Días transcurridos desde el último backup y umbral vigente."""
    conn = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"error": "DB no disponible"}), 500
        _ensure_backups_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("""
            SELECT generado_en, generado_por, periodo_desde, periodo_hasta,
                   cliente_nombre, propiedad_nombre, total_registros
            FROM backups_realizados
            ORDER BY generado_en DESC
            LIMIT 1
        """)
        row = cur.fetchone()
        cur.close()
        if not row:
            return jsonify({"ultimo": None, "dias": None})
        ultimo = dict(row)
        dias = (datetime.now(ultimo['generado_en'].tzinfo) - ultimo['generado_en']).days
        return jsonify({
            "ultimo": {k: (v.isoformat() if hasattr(v, 'isoformat') else v)
                       for k, v in ultimo.items()},
            "dias": dias,
        })
    except Exception as e:
        app_logger.error(f"backup_estado error: {e}", exc_info=True)
        return jsonify({"error": "Error interno"}), 500
    finally:
        if conn:
            conn.close()


@viewer_bp.route('/api/saved-filters', methods=['GET'])
@jwt_required()
@admin_required
def list_saved_filters():
    user_email = get_jwt_identity()
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=extras.RealDictCursor)
        cur.execute(
            "SELECT id, name, filters, updated_at FROM saved_filters WHERE user_email = %s ORDER BY name ASC",
            (user_email,)
        )
        rows = cur.fetchall()
        result = [{'id': r['id'], 'name': r['name'], 'filters': r['filters']} for r in rows]
        return jsonify({'saved_filters': result})
    except Exception as e:
        app_logger.error(f"Error listing saved filters: {e}", exc_info=True)
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@viewer_bp.route('/api/saved-filters', methods=['POST'])
@jwt_required()
@admin_required
def create_saved_filter():
    user_email = get_jwt_identity()
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    filters = data.get('filters')
    if not name:
        return jsonify({'error': 'El nombre es requerido.'}), 400
    if len(name) > 100:
        return jsonify({'error': 'El nombre no puede superar 100 caracteres.'}), 400
    if not isinstance(filters, dict):
        return jsonify({'error': 'Filtros inválidos.'}), 400
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=extras.RealDictCursor)
        cur.execute(
            """INSERT INTO saved_filters (user_email, name, filters)
               VALUES (%s, %s, %s)
               RETURNING id, name, filters""",
            (user_email, name, extras.Json(filters))
        )
        row = cur.fetchone()
        conn.commit()
        return jsonify({'saved_filter': {'id': row['id'], 'name': row['name'], 'filters': row['filters']}}), 201
    except pg_errors.UniqueViolation:
        if conn:
            conn.rollback()
        return jsonify({'error': f'Ya existe un filtro con el nombre "{name}".'}), 409
    except Exception as e:
        if conn:
            conn.rollback()
        app_logger.error("Error creating saved filter: %s", e, exc_info=True)
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@viewer_bp.route('/api/saved-filters/<int:filter_id>', methods=['PUT'])
@jwt_required()
@admin_required
def update_saved_filter(filter_id):
    user_email = get_jwt_identity()
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip() or None
    filters = data.get('filters')
    if name and len(name) > 100:
        return jsonify({'error': 'El nombre no puede superar 100 caracteres.'}), 400
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=extras.RealDictCursor)
        # Build dynamic SET clause
        sets, vals = [], []
        if name:
            sets.append("name = %s"); vals.append(name)
        if isinstance(filters, dict):
            sets.append("filters = %s"); vals.append(extras.Json(filters))
        if not sets:
            return jsonify({'error': 'Nada que actualizar.'}), 400
        sets.append("updated_at = NOW()")
        vals += [filter_id, user_email]
        cur.execute(
            f"UPDATE saved_filters SET {', '.join(sets)} WHERE id = %s AND user_email = %s RETURNING id, name, filters",
            vals
        )
        row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Filtro no encontrado.'}), 404
        conn.commit()
        return jsonify({'saved_filter': {'id': row['id'], 'name': row['name'], 'filters': row['filters']}})
    except pg_errors.UniqueViolation:
        if conn:
            conn.rollback()
        return jsonify({'error': 'Ya existe un filtro con ese nombre.'}), 409
    except Exception as e:
        if conn:
            conn.rollback()
        app_logger.error("Error updating saved filter: %s", e, exc_info=True)
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@viewer_bp.route('/api/saved-filters/<int:filter_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_saved_filter(filter_id):
    user_email = get_jwt_identity()
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM saved_filters WHERE id = %s AND user_email = %s RETURNING id",
            (filter_id, user_email)
        )
        if not cur.fetchone():
            return jsonify({'error': 'Filtro no encontrado.'}), 404
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn:
            conn.rollback()
        app_logger.error(f"Error deleting saved filter: {e}", exc_info=True)
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@viewer_bp.route('/logout')
def logout():
    try:
        app_logger.info("User logout requested")
        response = redirect('/')
        unset_jwt_cookies(response)
        app_logger.info("JWT cookies cleared, redirecting to login service")
        return response
    except Exception as e:
        app_logger.error(f"Error during logout: {e}", exc_info=True)
        # Fallback: just redirect without cookie clearing if there's an error
        return redirect('/')

# Viewer routes initialized
